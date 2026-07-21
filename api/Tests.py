import io

from django.test import SimpleTestCase
from django.db import IntegrityError
from openpyxl import load_workbook

from .utils import custom_exception_handler
from .Serializers import ManiobraSerializer
from .views import _generar_pdf_cta_port, _TEMPLATE_PATH

# Create your tests here.
class ErrorHandlingTests(SimpleTestCase):
	def test_integrity_error_reports_duplicate_field(self):
		exc = IntegrityError('duplicate key value violates unique constraint "maniobras_codigo_pis_key"\nDETAIL: Key (codigo_pis)=(ABC123) already exists.')

		response = custom_exception_handler(exc, context={})

		self.assertEqual(response.status_code, 400)
		self.assertEqual(response.data['detail']['codigo_pis'][0], 'Ya existe un registro con el valor "ABC123".')

	def test_serializer_reports_specific_codigo_pis_error(self):
		serializer = ManiobraSerializer(data={'codigo_pis': 'ABC 123'})

		self.assertFalse(serializer.is_valid())
		self.assertIn('codigo_pis', serializer.errors)
		self.assertEqual(str(serializer.errors['codigo_pis'][0]), 'Código PIS inválido.')


class CtaPortTipoServicioTests(SimpleTestCase):
	"""El tipo de servicio ya NO se adivina del texto: lo manda el botón
	'Tipo de Servicio'. Estos casos cubren la resolución del tipo, el dedup del
	segundo par (A19/B19) y el fallback de los registros viejos."""

	def _celdas(self, **campos):
		datos = {'folio': 'TEST-1', 'formato': 'excel'}
		datos.update(campos)
		resp = _generar_pdf_cta_port(datos, _TEMPLATE_PATH, 'test.pdf', formato='excel')
		self.assertEqual(resp.status_code, 200, getattr(resp, 'data', None))
		ws = load_workbook(io.BytesIO(resp.content))['CTA PORT FRABA CONTAINER']
		return ws

	def test_full_segundo_par_igual_no_se_escribe(self):
		ws = self._celdas(tipo_servicio='full', tipo='20 - 20 / DC - DC',
		                  contenedor='CBHU4284545 - TLLU3548653')
		self.assertEqual(ws['A17'].value, 2)
		self.assertEqual(ws['A18'].value, '20')
		self.assertEqual(ws['B18'].value, 'DC')
		self.assertIsNone(ws['A19'].value)
		self.assertIsNone(ws['B19'].value)
		self.assertEqual(ws['C17'].value, 'CBHU4284545 - TLLU3548653')

	def test_full_segundo_par_distinto_se_escribe(self):
		ws = self._celdas(tipo_servicio='full', tipo='20 - 40 / DC - HC',
		                  contenedor='CBHU4284545 - TLLU3548653')
		self.assertEqual(ws['A17'].value, 2)
		self.assertEqual(ws['A19'].value, '40')
		self.assertEqual(ws['B19'].value, 'HC')

	def test_sencillo_con_contenedor_largo_sigue_siendo_sencillo(self):
		# La heurística vieja (>12 caracteres) habría dicho full: ya no manda.
		ws = self._celdas(tipo_servicio='sencillo', tipo='20 / DC',
		                  contenedor='CBHU42845451234')
		self.assertEqual(ws['A17'].value, 1)
		self.assertEqual(ws['B17'].value, 'CONTENEDOR')

	def test_registro_viejo_sin_tipo_servicio_usa_heuristica(self):
		ws = self._celdas(tipo='20 / DC', contenedor='CBHU4284545-TLLU3548653')
		self.assertEqual(ws['A17'].value, 2)
		self.assertEqual(ws['B17'].value, 'CONTENEDORES')

	def test_carga_suelta_conserva_sus_dos_pares(self):
		ws = self._celdas(tipo_servicio='carga_suelta',
		                  tipo='9 - 14 / PALLETS - CARTONES',
		                  contenedor='CARGA SUELTA')
		self.assertEqual(ws['A17'].value, '9')
		self.assertEqual(ws['B17'].value, 'PALLETS')
		self.assertEqual(ws['A18'].value, '14')
		self.assertEqual(ws['B18'].value, 'CARTONES')
		self.assertEqual(ws['C17'].value, 'CARGA SUELTA')
