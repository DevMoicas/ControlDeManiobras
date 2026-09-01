"""El reporte de vacios pendientes de un coordinador.

Se pide en `formato: 'excel'` a proposito: ese camino devuelve el .xlsx tal cual
y se salta LibreOffice, que no esta instalado en el runner del CI —la conversion
a PDF la valida smoke_pdf.py dentro del contenedor—. Lo que se prueba aqui es lo
unico que puede decidir mal en silencio: A QUIEN se lleva cada fila.

Un reporte con el vacio de otro coordinador, o con uno ya entregado, manda a una
persona a buscar un contenedor que no le toca y nadie se entera hasta el patio.

Solo corre con:  Manage.py test api --settings=config.settings_test
"""
import io
from datetime import date

from django.contrib.auth import get_user_model
from django.db import connections
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from api.models import Maniobra, Vacio

URL = '/api/documentos/reporte-vacios/'
FILA_TITULOS = 5


class ReporteVaciosTests(TestCase):
    # Obligatorio en toda prueba con BD de este proyecto: RoleBasedRouter enruta
    # por el alias del hilo. Ver test_infra_bd.py.
    databases = {'default', 'standard'}

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        with connections['standard'].schema_editor() as editor:
            # `maniobras` primero aunque esta prueba no la use: desde la 0061
            # `vacios` tiene una FK a ella y no se puede crear sin su destino.
            editor.create_model(Maniobra)
            editor.create_model(Vacio)

    @classmethod
    def tearDownClass(cls):
        with connections['standard'].schema_editor() as editor:
            editor.delete_model(Vacio)
            editor.delete_model(Maniobra)
        super().tearDownClass()

    def setUp(self):
        self.usuario = get_user_model().objects.create_user('capturista', password='x')
        self.cliente = APIClient()
        self.cliente.force_authenticate(user=self.usuario)

    def pedir(self, coordinador, formato='excel'):
        return self.cliente.post(
            URL, {'coordinador': coordinador, 'formato': formato}, format='json')

    def filas_del_excel(self, respuesta):
        """Las filas de datos de la hoja devuelta, como listas de texto."""
        ws = load_workbook(io.BytesIO(b''.join(respuesta.streaming_content)
                                      if respuesta.streaming
                                      else respuesta.content)).active
        # La celda que se escribio vacia vuelve como None: se normaliza para que
        # las comprobaciones hablen del contenido y no de openpyxl.
        return [[c.value if c.value is not None else '' for c in fila]
                for fila in ws.iter_rows(min_row=FILA_TITULOS + 1)]

    def test_solo_los_pendientes_de_ese_coordinador(self):
        Vacio.objects.create(contenedor='SUYO1', status='pendiente',
                             coordinador='JORGE OCTAVIO MURAÑA')
        Vacio.objects.create(contenedor='ENTREGADO1', status='entregado',
                             coordinador='JORGE OCTAVIO MURAÑA')
        Vacio.objects.create(contenedor='DEOTRO1', status='pendiente',
                             coordinador='ANA LOPEZ')

        r = self.pedir('JORGE OCTAVIO MURAÑA')
        self.assertEqual(r.status_code, 200, getattr(r, 'data', r))
        contenedores = [fila[0] for fila in self.filas_del_excel(r)]
        self.assertEqual(contenedores, ['SUYO1'])

    def test_las_nueve_columnas_van_de_contenedor_a_op_del_viaje(self):
        Vacio.objects.create(
            contenedor='ABCD1234567', tipo_contenedor='40HC', patio='APM TERMINAL',
            fecha_maniobra=date(2026, 9, 12), fecha_entrega=date(2026, 9, 15),
            fecha_notificacion_cliente='AVISADO', status='pendiente',
            coordinador='ANA LOPEZ', operador='JUAN PEREZ',
            # Nada de esto sale en el reporte: Reprogramado, Fecha Reprogramacion
            # y Status EIR los quito el usuario para que lo demas quepa mas
            # grande, y Entrego / Cita / CD nunca estuvieron.
            reprogramado=True, fecha_reprogramacion=date(2026, 9, 20),
            status_eir='sin_eir_fisico',
            operador_entrega='NO SALE', cita='NO SALE', cd='NO SALE',
        )
        fila = self.filas_del_excel(self.pedir('ANA LOPEZ'))[0]
        self.assertEqual(fila, [
            'ABCD1234567', '40HC', 'APM TERMINAL', '12/09/2026', '15/09/2026',
            'AVISADO', 'Pendiente', 'ANA LOPEZ', 'JUAN PEREZ',
        ])
        # Ni rastro de lo que se quito, en ninguna columna.
        for fuera in ('Sí', '20/09/2026', 'Sin EIR Físico', 'NO SALE'):
            self.assertNotIn(fuera, fila)

    def test_las_etiquetas_son_las_de_la_pantalla_y_no_las_de_la_base(self):
        # En la base el status vive en minusculas; el selector lo pinta
        # 'Pendiente' y el papel tiene que decir lo mismo que la pantalla.
        Vacio.objects.create(contenedor='X1', status='pendiente',
                             coordinador='ANA LOPEZ')
        self.assertEqual(self.filas_del_excel(self.pedir('ANA LOPEZ'))[0][6], 'Pendiente')

    def test_una_mayuscula_de_mas_no_deja_el_reporte_vacio(self):
        # El coordinador se guarda como texto, no como FK.
        Vacio.objects.create(contenedor='X1', status='pendiente', coordinador='Ana Lopez')
        self.assertEqual(self.pedir('ANA LOPEZ').status_code, 200)

    def test_sin_coordinador_no_se_genera_nada(self):
        r = self.pedir('   ')
        self.assertEqual(r.status_code, 400)
        self.assertIn('coordinador', r.data['detail'].lower())

    def test_un_coordinador_sin_pendientes_lo_dice_en_vez_de_dar_un_pdf_vacio(self):
        Vacio.objects.create(contenedor='X1', status='entregado', coordinador='ANA LOPEZ')
        r = self.pedir('ANA LOPEZ')
        self.assertEqual(r.status_code, 400)
        self.assertIn('no tiene vacíos pendientes', r.data['detail'])

    def test_una_formula_en_un_campo_de_texto_no_se_ejecuta_en_el_excel(self):
        # CWE-1236: sin la comilla, Excel/LibreOffice lo trataria como formula.
        Vacio.objects.create(contenedor='=1+1', status='pendiente', coordinador='ANA LOPEZ')
        self.assertEqual(self.filas_del_excel(self.pedir('ANA LOPEZ'))[0][0], "'=1+1")
