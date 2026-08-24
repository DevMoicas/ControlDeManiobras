"""Pruebas del reporte de viaje de los coordinadores.

Se cubre lo que puede romper en silencio: que un folio no acabe con dos reportes,
que los Sí/No conserven sus TRES estados (sí / no / sin contestar), que los
campos calculados no mientan cuando falta un operando, y que las cinco cargas de
combustible se escriban en una sola petición sin pisarse entre ellas.

Ver docs/planes/PLAN_REPORTE_COORDINADORES.md (rama main).

Solo corre con:  Manage.py test api --settings=config.settings_test
"""
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from api.models import ReporteViaje, CargaCombustible, CARGAS_POR_REPORTE
from api.views import _TEMPLATE_REPORTE

URL = '/api/reportes-viaje/'


class BaseReporte(TestCase):
    # Obligatorio en toda prueba con BD de este proyecto: RoleBasedRouter enruta
    # por el alias del hilo. Ver test_infra_bd.py.
    databases = {'default', 'standard'}

    def setUp(self):
        self.usuario = get_user_model().objects.create_user('reporte_user', password='x')
        self.cliente = APIClient()
        self.cliente.force_authenticate(user=self.usuario)

    def crear(self, **campos):
        campos.setdefault('folio', 'F-2279')
        return self.cliente.post(URL, campos, format='json')


class ReporteViajeTests(BaseReporte):

    # ── Un reporte por folio ─────────────────────────────────────────────
    def test_se_crea_un_reporte_con_su_folio(self):
        r = self.crear(coordinador='Ali', cliente='YAZAKI')
        self.assertEqual(r.status_code, 201, r.data)
        self.assertEqual(ReporteViaje.objects.get().folio, 'F-2279')

    def test_un_segundo_reporte_del_mismo_folio_se_rechaza_con_mensaje_legible(self):
        """El apiClient del frontend solo lee la clave 'detail': bajo cualquier
        otra, el usuario vería un inútil "HTTP 400"."""
        self.crear()
        r = self.crear()
        self.assertEqual(r.status_code, 400)
        self.assertIn('ya tiene un reporte', str(r.data.get('detail', '')))
        self.assertEqual(ReporteViaje.objects.count(), 1)

    def test_el_folio_no_puede_quedar_vacio(self):
        for vacio in ('', '   '):
            r = self.crear(folio=vacio)
            self.assertEqual(r.status_code, 400, vacio)
            self.assertIn('no puede quedar vacío', str(r.data.get('detail', '')))

    def test_el_folio_se_guarda_sin_espacios_de_sobra(self):
        r = self.crear(folio='  F-2279  ')
        self.assertEqual(r.status_code, 201, r.data)
        self.assertEqual(ReporteViaje.objects.get().folio, 'F-2279')

    def test_editar_un_reporte_sin_cambiarle_el_folio_no_choca_consigo_mismo(self):
        rid = self.crear().data['id']
        r = self.cliente.patch(f'{URL}{rid}/', {'folio': 'F-2279', 'coordinador': 'Mari'},
                               format='json')
        self.assertEqual(r.status_code, 200, r.data)

    def test_filtrar_por_folio_dice_si_ya_hay_reporte(self):
        """La pantalla pregunta antes de abrir uno nuevo, en vez de descubrirlo
        con un 400 al guardar."""
        self.crear(folio='F-2279')
        self.crear(folio='R-2280')
        r = self.cliente.get(f'{URL}?folio=F-2279')
        self.assertEqual([x['folio'] for x in r.data['results']], ['F-2279'])

    # ── Los Sí/No conservan tres estados ─────────────────────────────────
    def test_los_si_no_admiten_si_no_y_sin_contestar(self):
        """null NO es False: es lo que deja el "SI / NO" impreso intacto en el
        Excel para rodearlo a mano. Colapsarlo haría que un reporte a medio
        llenar afirmara "No" en lo que nadie ha contestado."""
        rid = self.crear().data['id']
        for campo in ('reparacion', 'rescate', 'maniobra_vacio', 'estadias'):
            for valor in (True, False, None):
                r = self.cliente.patch(f'{URL}{rid}/', {campo: valor}, format='json')
                self.assertEqual(r.status_code, 200, r.data)
                self.assertIs(r.data[campo], valor, f'{campo} = {valor}')

    def test_un_reporte_recien_creado_trae_los_si_no_sin_contestar(self):
        r = self.crear()
        for campo in ('reparacion', 'rescate', 'maniobra_vacio', 'estadias'):
            self.assertIsNone(r.data[campo], campo)

    # ── Calculados ───────────────────────────────────────────────────────
    def test_km_totales_es_la_resta(self):
        r = self.crear(km_inicial=124500, km_final=125380)
        self.assertEqual(r.data['km_totales'], 880)

    def test_km_totales_es_none_mientras_falte_un_operando(self):
        for campos in ({'km_inicial': 124500}, {'km_final': 125380}, {}):
            ReporteViaje.objects.all().delete()
            r = self.crear(**campos)
            self.assertIsNone(r.data['km_totales'], campos)

    def test_el_total_de_una_carga_es_litros_por_precio(self):
        r = self.crear(cargas=[{'orden': 1, 'litros_diesel': '300', 'precio_litro': '24.80'}])
        self.assertEqual(r.data['cargas'][0]['total'], '7440.00')

    def test_el_total_de_una_carga_es_none_sin_precio(self):
        r = self.crear(cargas=[{'orden': 1, 'litros_diesel': '300'}])
        self.assertIsNone(r.data['cargas'][0]['total'])

    def test_el_total_de_urea_se_captura_no_se_calcula(self):
        """El papel no trae precio por litro para la urea, así que su total no
        es derivable: entra tal cual lo escribe el coordinador."""
        r = self.crear(cargas=[{'orden': 1, 'litros_urea': '20', 'total_urea': '496.00'}])
        self.assertEqual(r.data['cargas'][0]['total_urea'], '496.00')

    def test_el_rendimiento_suma_el_diesel_de_todas_las_cargas(self):
        """El rendimiento es del viaje entero, no de una parada: 880 km entre los
        300 + 100 litros de las dos cargas."""
        r = self.crear(
            km_inicial=124500, km_final=125380,
            cargas=[{'orden': 1, 'litros_diesel': '300', 'precio_litro': '24.80'},
                    {'orden': 2, 'litros_diesel': '100', 'precio_litro': '24.50'}],
        )
        self.assertEqual(r.data['rendimiento'], '2.20')

    def test_el_rendimiento_es_none_sin_km_o_sin_litros(self):
        r = self.crear(cargas=[{'orden': 1, 'litros_diesel': '300'}])
        self.assertIsNone(r.data['rendimiento'])

        ReporteViaje.objects.all().delete()
        r = self.crear(km_inicial=1, km_final=2)
        self.assertIsNone(r.data['rendimiento'])

    # ── Las cinco cargas ─────────────────────────────────────────────────
    def test_las_cinco_cargas_entran_en_una_sola_peticion(self):
        cargas = [{'orden': i, 'litros_diesel': f'{i}00'} for i in range(1, 6)]
        r = self.crear(cargas=cargas)
        self.assertEqual(r.status_code, 201, r.data)
        self.assertEqual(CargaCombustible.objects.count(), 5)
        self.assertEqual([c['orden'] for c in r.data['cargas']], [1, 2, 3, 4, 5])

    def test_reenviar_un_renglon_lo_pisa_y_no_toca_los_demas(self):
        """Upsert por `orden`. El coordinador captura una parada hoy y otra
        pasado mañana: la segunda no puede borrar la primera."""
        rid = self.crear(cargas=[{'orden': 1, 'litros_diesel': '300'},
                                 {'orden': 2, 'litros_diesel': '100'}]).data['id']
        r = self.cliente.patch(f'{URL}{rid}/',
                               {'cargas': [{'orden': 2, 'litros_diesel': '150'}]},
                               format='json')
        self.assertEqual(r.status_code, 200, r.data)
        self.assertEqual(CargaCombustible.objects.count(), 2)
        por_orden = {c['orden']: c['litros_diesel'] for c in r.data['cargas']}
        self.assertEqual(por_orden, {1: '300.00', 2: '150.00'})

    def test_un_patch_de_otro_campo_no_borra_las_cargas(self):
        """`cargas` ausente no es `cargas: []`. Sin esa distinción, guardar el
        coordinador se llevaría por delante todo el combustible."""
        rid = self.crear(cargas=[{'orden': 1, 'litros_diesel': '300'}]).data['id']
        self.cliente.patch(f'{URL}{rid}/', {'coordinador': 'Ali'}, format='json')
        self.assertEqual(CargaCombustible.objects.count(), 1)

    def test_un_renglon_fuera_de_rango_se_rechaza_con_400_y_no_con_500(self):
        """El CHECK de la base lo rechaza igual, pero como IntegrityError, que el
        usuario ve como un 500 sin explicación."""
        for fuera in (0, CARGAS_POR_REPORTE + 1):
            ReporteViaje.objects.all().delete()
            r = self.crear(cargas=[{'orden': fuera, 'litros_diesel': '300'}])
            self.assertEqual(r.status_code, 400, f'orden {fuera}: {r.data}')

    # ── Permisos ─────────────────────────────────────────────────────────
    def test_cualquier_usuario_autenticado_crea_y_edita(self):
        """Sin candado de rol: lo llena y lo ve cualquiera (usuario, 2026-08-24)."""
        self.assertFalse(self.usuario.is_staff)
        rid = self.crear().data['id']
        self.assertEqual(
            self.cliente.patch(f'{URL}{rid}/', {'coordinador': 'Ali'}, format='json').status_code,
            200)

    def test_borrar_un_reporte_se_reserva_a_admin(self):
        """Un reporte es un documento que se firma: borrarlo no es corregir."""
        rid = self.crear().data['id']
        self.assertEqual(self.cliente.delete(f'{URL}{rid}/').status_code, 403)
        self.assertEqual(ReporteViaje.objects.count(), 1)

        self.usuario.is_staff = True
        self.usuario.save()
        self.assertEqual(self.cliente.delete(f'{URL}{rid}/').status_code, 204)

    def test_sin_autenticar_no_se_ve_nada(self):
        self.assertIn(APIClient().get(URL).status_code, (401, 403))

    def test_borrar_el_reporte_se_lleva_sus_cargas(self):
        rid = self.crear(cargas=[{'orden': 1, 'litros_diesel': '300'}]).data['id']
        self.usuario.is_staff = True
        self.usuario.save()
        self.cliente.delete(f'{URL}{rid}/')
        self.assertEqual(CargaCombustible.objects.count(), 0)


class ReporteViajeDocumentoTests(BaseReporte):
    """El Excel que sale de la plantilla del papel.

    Solo se prueba `formato=excel`: es el que ejercita TODO el llenado. El PDF
    es ese mismo archivo pasado por LibreOffice, que no está en el entorno de
    pruebas y ya cubren las cartas porte.
    """

    def descargar(self, rid, formato='excel'):
        return self.cliente.get(f'{URL}{rid}/documento/?formato={formato}')

    @staticmethod
    def hoja(respuesta):
        return load_workbook(BytesIO(respuesta.content)).worksheets[0]

    @staticmethod
    def plantilla():
        """La plantilla sin tocar, para comparar lo que NO debe cambiar."""
        return load_workbook(str(_TEMPLATE_REPORTE)).worksheets[0]

    def test_el_excel_trae_lo_capturado_en_sus_celdas(self):
        rid = self.crear(
            fecha='2026-08-24', coordinador='Ali', cliente='YAZAKI',
            origen='Manzanillo', destino='Guadalajara', operador='Juan Perez',
            unidad='93-AF-2K', remolque_1='R-101', remolque_2='R-102',
            km_inicial=124500, km_final=125380,
            cargas=[{'orden': 1, 'litros_diesel': '300', 'precio_litro': '24.80',
                     'litros_urea': '20', 'total_urea': '496.00'}],
        ).data['id']
        r = self.descargar(rid)
        self.assertEqual(r.status_code, 200, r.content[:200])
        ws = self.hoja(r)

        self.assertEqual(ws['C2'].value, '24/08/2026')
        self.assertEqual(ws['G2'].value, 'ALI')          # todo va en mayúsculas
        self.assertEqual(ws['J2'].value, 'F-2279')
        self.assertEqual(ws['E3'].value, 'YAZAKI')
        self.assertEqual(ws['B4'].value, 'MANZANILLO')
        self.assertEqual(ws['E4'].value, 'GUADALAJARA')
        # UNIDAD Y PORTAS: tracto y remolques en una sola celda
        self.assertEqual(ws['D8'].value, '93-AF-2K / R-101, R-102')
        self.assertEqual(ws['H8'].value, 124500)
        self.assertEqual(ws['H9'].value, 125380)
        # Renglón 1 de EN TRAYECTO = fila 12
        # float y no Decimal: openpyxl los guarda asi porque Excel no tiene
        # tipo decimal. Lo que importa es que sean NUMEROS y no texto.
        self.assertEqual(ws['C12'].value, 300.0)
        self.assertEqual(ws['E12'].value, 24.8)
        self.assertEqual(ws['K12'].value, 496.0)

    def test_los_calculados_van_como_numero_y_no_como_texto(self):
        """Si salieran como cadena, el Excel no los sumaría ni los formatearía."""
        rid = self.crear(
            km_inicial=124500, km_final=125380,
            cargas=[{'orden': 1, 'litros_diesel': '300', 'precio_litro': '24.80'}],
        ).data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['K8'].value, 880)         # KM TOTALES
        self.assertEqual(ws['K9'].value, 2.93)        # RENDIMIENTO
        self.assertEqual(ws['G12'].value, 7440.0)     # TOTAL del renglón
        # Lo que de verdad se comprueba: que NINGUNO sea texto.
        for celda in ('K8', 'K9', 'G12'):
            self.assertNotIsInstance(ws[celda].value, str, celda)

    def test_las_horas_salen_en_la_hora_de_operacion_y_no_en_utc(self):
        """Se guardan en UTC y el documento lo arma el servidor: sin convertir,
        el papel saldría seis horas corrido."""
        rid = self.crear(cita='2026-08-24T14:00:00Z').data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['D5'].value, '24/08/2026 08:00')

    # ── La regla de los Sí/No ────────────────────────────────────────────
    def test_un_si_no_contestado_pisa_la_etiqueta(self):
        rid = self.crear(reparacion=True, rescate=False, maniobra_vacio=True).data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['G17'].value, 'SI')   # ¿REPARACIÓN?
        self.assertEqual(ws['C18'].value, 'NO')   # ¿RESCATE?
        self.assertEqual(ws['J20'].value, 'SI')   # MANIOBRA DE VACÍO

    def test_un_si_no_sin_contestar_deja_la_etiqueta_para_rodearla_a_mano(self):
        rid = self.crear().data['id']
        ws = self.hoja(self.descargar(rid))
        original = self.plantilla()
        for celda in ('G17', 'C18', 'J20'):
            self.assertEqual(ws[celda].value, original[celda].value, celda)

    def test_las_estadias_borran_la_palabra_que_no_aplica(self):
        """Aquí el SI y el NO son dos celdas distintas, no una etiqueta."""
        rid = self.crear(estadias=True).data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['E22'].value, 'SI')
        self.assertIn(ws['F22'].value, (None, ''))

    def test_la_recoleccion_se_marca_con_una_X_en_su_casilla(self):
        """Las dos palabras se quedan impresas, como en el papel: lo que cambia
        es la casilla de al lado."""
        rid = self.crear(recoleccion='propio').data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['J3'].value, 'X')          # casilla de PROPIO
        self.assertIn(ws['L3'].value, (None, ''))      # la de TERCERO, vacía
        self.assertEqual(ws['I3'].value, 'PROPIO')
        self.assertEqual(ws['K3'].value, 'TERCERO')

    def test_tercero_marca_la_otra_casilla(self):
        rid = self.crear(recoleccion='tercero').data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['L3'].value, 'X')
        self.assertIn(ws['J3'].value, (None, ''))

    def test_sin_respuesta_no_se_marca_ninguna_casilla(self):
        rid = self.crear().data['id']
        ws = self.hoja(self.descargar(rid))
        original = self.plantilla()
        for celda in ('I3', 'K3', 'E22', 'F22'):
            self.assertEqual(ws[celda].value, original[celda].value, celda)
        for casilla in ('J3', 'L3'):
            self.assertIn(ws[casilla].value, (None, ''), casilla)

    def test_un_nombre_largo_normal_sale_a_tamano_completo_en_dos_lineas(self):
        """La plantilla le dio a la casilla alto para dos lineas (fila 4 a 30.75).
        Con el wrap_text que ya trae, un nombre completo cabe entero sin encoger
        la letra — que es lo que se lee bien en el papel."""
        largo = 'JORGE OCTAVIO MURANA MANZO'          # 26 caracteres
        rid = self.crear(operador=largo).data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['I4'].value, largo)
        self.assertTrue(ws['I4'].alignment.wrap_text)
        self.assertFalse(ws['I4'].alignment.shrinkToFit)

    def test_solo_lo_que_no_cabe_ni_en_dos_lineas_se_encoge(self):
        """Una celda COMBINADA no autoajusta su alto ni en Excel ni en
        LibreOffice, asi que pasado ese punto el texto se cortaria en silencio.
        Ahi si se encoge la letra, y wrap_text tiene que apagarse porque los dos
        son incompatibles."""
        rid = self.crear(operador='A' * 80).data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertTrue(ws['I4'].alignment.shrinkToFit)
        self.assertFalse(ws['I4'].alignment.wrap_text)

    def test_encoger_no_se_come_la_casilla_de_al_lado(self):
        """K4:L4 es OTRA casilla, con su propia linea divisoria. Encoger el texto
        la respeta; ampliar la combinacion la habria borrado."""
        rid = self.crear(operador='UN NOMBRE MUY LARGO QUE NO CABE NI DE BROMA JAMAS').data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertIn('I4:J4', [str(m) for m in ws.merged_cells.ranges])
        self.assertIn('K4:L4', [str(m) for m in ws.merged_cells.ranges])

    # ── Impresión ────────────────────────────────────────────────────────
    def test_el_documento_sale_en_una_sola_hoja_horizontal(self):
        """Son 12 columnas anchas: en vertical LibreOffice parte el formato y
        saca una segunda página con el pico derecho, ilegible."""
        rid = self.crear().data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws.page_setup.orientation, 'landscape')
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        self.assertEqual(ws.page_setup.fitToHeight, 1)
        # Sin fitToPage, los dos de arriba no hacen nada.
        self.assertTrue(ws.sheet_properties.pageSetUpPr.fitToPage)
        # El doble alto de la fila del operador lo pone el codigo, no el .xlsx:
        # en el archivo se pierde en cuanto alguien lo reedita o lo reemplaza.
        self.assertEqual(ws.row_dimensions[4].height, 33.0)
        # openpyxl lo normaliza a referencia absoluta con el nombre de la hoja.
        self.assertIn('$A$1:$L$25', str(ws.print_area))

    # ── El espacio de la firma no se toca ────────────────────────────────
    def test_el_espacio_de_la_firma_se_queda_vacio(self):
        """La FIRMA DEL COORDINADOR va en papel."""
        rid = self.crear(comentarios='sin novedad').data['id']
        ws = self.hoja(self.descargar(rid))
        self.assertEqual(ws['A24'].value, 'SIN NOVEDAD')
        self.assertIsNone(ws['F25'].value)

    def test_el_documento_exige_autenticacion(self):
        rid = self.crear().data['id']
        self.assertIn(APIClient().get(f'{URL}{rid}/documento/').status_code, (401, 403))


class ReporteViajeCampanaOpcionalTests(BaseReporte):
    """Nada es obligatorio salvo el folio.

    El reporte se llena por etapas: hay viajes que no usan la mitad del formato y
    datos que no se saben hasta días después. Guardar a medias y completar luego
    tiene que funcionar (usuario, 2026-08-24).
    """

    # Todos los campos de texto del modelo, derivados igual que en el serializer.
    TEXTOS = ('coordinador', 'servicio', 'cliente', 'recoleccion', 'origen',
              'destino', 'operador', 'unidad', 'remolque_1', 'remolque_2',
              'reparacion_que', 'rescate_unidad', 'rescate_operador',
              'patio_entrega', 'unidad_vacio', 'operador_vacio', 'comentarios')

    def test_con_solo_el_folio_ya_se_guarda(self):
        r = self.crear()
        self.assertEqual(r.status_code, 201, r.data)

    def test_un_null_en_cualquier_campo_de_texto_se_guarda_como_vacio(self):
        """El apiClient del frontend convierte "" en null (sanitizarPayload), así
        que un formulario a medio llenar llega entero en nulls. Sin esto son 17
        errores de "This field may not be null" y no se puede guardar nada."""
        r = self.crear(**{campo: None for campo in self.TEXTOS})
        self.assertEqual(r.status_code, 201, r.data)
        for campo in self.TEXTOS:
            self.assertEqual(r.data[campo], '', campo)

    def test_un_null_en_los_numericos_y_las_fechas_tambien_pasa(self):
        r = self.crear(
            km_inicial=None, km_final=None, litros_aceite=None, precio_aceite=None,
            reparacion_costo=None, estadias_horas=None,
            fecha=None, cita=None, salida_puerto=None, inicio_pactado=None,
            salida_real=None, llegada_cliente=None, descarga=None,
            llegada_manzanillo=None, cita_vacio=None,
        )
        self.assertEqual(r.status_code, 201, r.data)
        self.assertIsNone(r.data['km_inicial'])
        self.assertIsNone(r.data['cita'])

    def test_vaciar_un_campo_ya_escrito_no_da_error(self):
        """Borrar lo que se capturó por error tiene que poder hacerse."""
        rid = self.crear(coordinador='Ali', comentarios='algo').data['id']
        r = self.cliente.patch(f'{URL}{rid}/',
                               {'coordinador': None, 'comentarios': None}, format='json')
        self.assertEqual(r.status_code, 200, r.data)
        self.assertEqual(r.data['coordinador'], '')
        self.assertEqual(r.data['comentarios'], '')

    def test_el_reporte_se_completa_por_etapas(self):
        """Cada tanda de datos entra sin tocar lo anterior: la cita hoy, el
        combustible en carretera, el regreso dentro de tres días."""
        rid = self.crear(coordinador='Ali', cita='2026-08-24T14:00:00Z').data['id']

        self.cliente.patch(f'{URL}{rid}/', {'km_inicial': 124500}, format='json')
        self.cliente.patch(f'{URL}{rid}/',
                           {'cargas': [{'orden': 1, 'litros_diesel': '300'}]}, format='json')
        r = self.cliente.patch(f'{URL}{rid}/',
                               {'km_final': 125380, 'estadias': True,
                                'patio_entrega': 'Patio Norte'}, format='json')

        self.assertEqual(r.status_code, 200, r.data)
        self.assertEqual(r.data['coordinador'], 'Ali')          # lo de la primera tanda
        self.assertEqual(r.data['km_inicial'], 124500)          # lo de la segunda
        self.assertEqual(len(r.data['cargas']), 1)              # lo de la tercera
        self.assertEqual(r.data['km_totales'], 880)             # y el calculado al día

    def test_el_folio_sigue_siendo_lo_unico_obligatorio(self):
        """null en el folio no puede colarse como cadena vacía y crear un reporte
        sin viaje: se rechaza con el mismo mensaje legible que la cadena vacía."""
        r = self.crear(folio=None)
        self.assertEqual(r.status_code, 400)
        self.assertIn('no puede quedar vacío', str(r.data.get('detail', '')))


class ReporteViajeRendimientoGuardadoTests(BaseReporte):
    """El rendimiento es la excepcion: se GUARDA en columna.

    Lo piden los informes que vendran despues, que agregan rendimientos de
    muchos viajes y no pueden recalcularlos arrastrando las cargas de cada uno
    (decision del usuario, 2026-08-24). El riesgo de que se desfase de sus
    operandos se acota recalculandolo en CADA escritura.
    """

    def crear_con_viaje(self, **extra):
        campos = dict(km_inicial=124500, km_final=125380,
                      cargas=[{'orden': 1, 'litros_diesel': '300',
                               'precio_litro': '24.80'}])
        campos.update(extra)
        return self.crear(**campos)

    def test_se_guarda_en_la_columna_no_solo_en_la_respuesta(self):
        rid = self.crear_con_viaje().data['id']
        self.assertEqual(ReporteViaje.objects.get(pk=rid).rendimiento,
                         Decimal('2.93'))          # 880 km / 300 lt

    def test_se_recalcula_al_corregir_el_kilometraje(self):
        """Si el kilometraje cambia y el rendimiento no, la cifra guardada
        empieza a mentir. Por eso se refresca en cada escritura."""
        rid = self.crear_con_viaje().data['id']
        r = self.cliente.patch(f'{URL}{rid}/', {'km_final': 125100}, format='json')
        self.assertEqual(r.status_code, 200, r.data)
        # 600 km / 300 lt
        self.assertEqual(ReporteViaje.objects.get(pk=rid).rendimiento, Decimal('2.00'))

    def test_se_recalcula_al_anadir_otra_carga(self):
        rid = self.crear_con_viaje().data['id']
        self.cliente.patch(f'{URL}{rid}/',
                           {'cargas': [{'orden': 2, 'litros_diesel': '100'}]},
                           format='json')
        # 880 km / (300 + 100) lt
        self.assertEqual(ReporteViaje.objects.get(pk=rid).rendimiento, Decimal('2.20'))

    def test_sin_operandos_se_guarda_vacio_y_no_cero(self):
        """Un 0.00 guardado se leeria como "este camion no rinde nada", que es
        una afirmacion; NULL dice que no se sabe."""
        rid = self.crear().data['id']
        self.assertIsNone(ReporteViaje.objects.get(pk=rid).rendimiento)

    def test_deja_de_haber_rendimiento_si_se_borra_el_kilometraje(self):
        rid = self.crear_con_viaje().data['id']
        self.cliente.patch(f'{URL}{rid}/', {'km_final': None}, format='json')
        self.assertIsNone(ReporteViaje.objects.get(pk=rid).rendimiento)

    def test_no_se_puede_mandar_desde_fuera(self):
        """read_only: el frontend manda el reporte entero, y ahi viaja su copia
        calculada en vivo. Si se aceptara, un cliente podria guardar cualquier
        cifra sin relacion con los kilometros ni los litros."""
        rid = self.crear_con_viaje(rendimiento='999.99').data['id']
        self.assertEqual(ReporteViaje.objects.get(pk=rid).rendimiento, Decimal('2.93'))

    def test_la_urea_no_cuenta_en_el_rendimiento(self):
        rid = self.crear_con_viaje(
            cargas=[{'orden': 1, 'litros_diesel': '300', 'precio_litro': '24.80',
                     'litros_urea': '500', 'total_urea': '100'}]).data['id']
        self.assertEqual(ReporteViaje.objects.get(pk=rid).rendimiento, Decimal('2.93'))
