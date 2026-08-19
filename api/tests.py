import io
import logging
from types import SimpleNamespace

from django.test import SimpleTestCase
from django.db import IntegrityError
from django.contrib.auth.models import AnonymousUser
from openpyxl import load_workbook

from .utils import custom_exception_handler
from .Serializers import ManiobraSerializer
from .views import _generar_pdf_cta_port, _resolver_cliente, _TEMPLATE_PATH

# Create your tests here.
class ErrorHandlingTests(SimpleTestCase):
	def test_integrity_error_reports_duplicate_field(self):
		exc = IntegrityError('duplicate key value violates unique constraint "maniobras_codigo_pis_key"\nDETAIL: Key (codigo_pis)=(ABC123) already exists.')

		response = custom_exception_handler(exc, context={})

		self.assertEqual(response.status_code, 400)
		self.assertEqual(response.data['detail']['codigo_pis'][0], 'Ya existe un registro con el valor "ABC123".')

	# El validador de formato de codigo_pis se eliminó a propósito (ver el
	# comentario `ponytail:` en Serializers.py): ahora acepta texto libre y el
	# max_length=100 del modelo es lo que evita el error de columna. Estos dos
	# casos fijan lo que el serializer valida HOY.
	def test_solicita_es_el_unico_campo_obligatorio(self):
		serializer = ManiobraSerializer(data={'codigo_pis': 'ABC 123'})

		self.assertFalse(serializer.is_valid())
		self.assertIn('solicita', serializer.errors)
		self.assertNotIn('codigo_pis', serializer.errors)

	def test_codigo_pis_acepta_texto_libre(self):
		serializer = ManiobraSerializer(data={'solicita': 'JUAN', 'codigo_pis': 'ABC 123'})

		self.assertTrue(serializer.is_valid(), serializer.errors)

	def test_solicita_respeta_su_limite_de_longitud(self):
		serializer = ManiobraSerializer(data={'solicita': 'X' * 31})

		self.assertFalse(serializer.is_valid())
		self.assertIn('solicita', serializer.errors)


class ResolucionClientePorFolioTests(SimpleTestCase):
	"""Al elegir un folio, la Carta Porte tiene que traer SU cliente aunque haya
	homónimos: dos "YAZAKI" con distinta dirección solo se distinguen por el FK.
	Si alguien vuelve a resolverlo por nombre, estos casos fallan.
	Sin BD a propósito: `maniobras` es managed=False y el runner no crea la tabla."""

	def _cliente(self, ciudad):
		return SimpleNamespace(nombre_cliente='YAZAKI', domicilio='', colonia='', ciudad=ciudad)

	def test_el_fk_gana_sobre_el_homonimo_del_mapa_por_nombre(self):
		silao, durango = self._cliente('Silao'), self._cliente('Durango')
		maniobra = SimpleNamespace(cliente='YAZAKI', cliente_fk_id=2, cliente_fk=silao)

		self.assertIs(_resolver_cliente(maniobra, {'YAZAKI': durango}), silao)

	def test_folio_viejo_sin_fk_cae_al_mapa_por_nombre(self):
		durango = self._cliente('Durango')
		maniobra = SimpleNamespace(cliente='YAZAKI', cliente_fk_id=None, cliente_fk=None)

		self.assertIs(_resolver_cliente(maniobra, {'YAZAKI': durango}), durango)

	def test_el_serializer_expone_cliente_id_apuntando_al_fk(self):
		campos = ManiobraSerializer().fields

		self.assertEqual(campos['cliente_id'].source, 'cliente_fk')
		self.assertNotIn('cliente_fk', campos)


class LoggingSeguridadTests(SimpleTestCase):
	"""3.4.2 — los eventos de seguridad tienen que llegar al logger 'api.security'.
	Si alguien desconecta una señal o sube el nivel del handler, esto falla."""

	def test_acceso_denegado_403_se_loguea(self):
		from rest_framework.exceptions import PermissionDenied
		from django.test import RequestFactory

		request = RequestFactory().delete('/api/maniobras/7/')
		request.user = AnonymousUser()

		with self.assertLogs('api.security', level='WARNING') as log:
			custom_exception_handler(PermissionDenied(), {'request': request})
		self.assertIn('acceso denegado 403', log.output[0])
		self.assertIn('/api/maniobras/7/', log.output[0])

	def test_error_400_no_ensucia_el_log_de_seguridad(self):
		from rest_framework.exceptions import ValidationError as DRFValidationError
		from django.test import RequestFactory

		request = RequestFactory().post('/api/maniobras/')
		request.user = AnonymousUser()

		with self.assertNoLogs('api.security', level='WARNING'):
			custom_exception_handler(DRFValidationError({'solicita': ['requerido']}),
			                         {'request': request})

	# Emitir las señales de verdad arrastraría a los receptores propios de axes,
	# que consultan la BD. Se comprueba en dos pasos: que el nuestro está
	# registrado (por dispatch_uid) y que hace lo suyo al ejecutarse.
	def _uids_conectados(self, signal):
		return [entrada[0][0] for entrada in signal.receivers]

	def test_senal_de_login_fallido_esta_conectada(self):
		from django.contrib.auth.signals import user_login_failed
		from django.test import RequestFactory
		from .Apps import _log_login_fallido

		self.assertIn('api_login_fallido', self._uids_conectados(user_login_failed))

		with self.assertLogs('api.security', level='WARNING') as log:
			_log_login_fallido(sender=None,
			                   credentials={'username': 'atacante'},
			                   request=RequestFactory().post('/api/token/'))
		self.assertIn('login FALLIDO', log.output[0])
		self.assertIn('atacante', log.output[0])

	def test_senal_de_lockout_esta_conectada(self):
		from axes.signals import user_locked_out
		from django.test import RequestFactory
		from .Apps import _log_lockout

		self.assertIn('api_lockout', self._uids_conectados(user_locked_out))

		with self.assertLogs('api.security', level='WARNING') as log:
			_log_lockout(sender=None,
			             request=RequestFactory().post('/api/token/'),
			             username='atacante')
		self.assertIn('LOCKOUT', log.output[0])
		self.assertIn('atacante', log.output[0])


class MfaAdminTests(SimpleTestCase):
	"""3.3.5 — /admin exige un dispositivo OTP verificado, no solo is_staff.
	Si alguien quita el intercambio de clase en config/urls.py, esto falla."""

	class _UsuarioFalso:
		is_active = True
		is_staff = True

		def __init__(self, verificado):
			self._verificado = verificado

		def is_verified(self):
			return self._verificado

	def _admin_site(self):
		import config.urls  # noqa: F401  — importarlo es lo que hace el intercambio
		from django.contrib import admin
		return admin.site

	def test_el_sitio_admin_es_un_otpadminsite(self):
		from django_otp.admin import OTPAdminSite
		self.assertIsInstance(self._admin_site(), OTPAdminSite)

	def test_staff_sin_otp_verificado_no_entra(self):
		from django.test import RequestFactory

		req = RequestFactory().get('/admin/')
		req.user = self._UsuarioFalso(verificado=False)
		self.assertFalse(self._admin_site().has_permission(req))

	def test_staff_con_otp_verificado_si_entra(self):
		from django.test import RequestFactory

		req = RequestFactory().get('/admin/')
		req.user = self._UsuarioFalso(verificado=True)
		self.assertTrue(self._admin_site().has_permission(req))


class CspPorRutaTests(SimpleTestCase):
	"""La API va con default-src 'none' (solo devuelve JSON), pero /admin/ es HTML
	de Django con su propio CSS/JS/imágenes: con 'none' el panel sale sin estilos
	y el navegador bloquea el QR de alta del dispositivo TOTP."""

	def _csp(self, path):
		from django.test import RequestFactory
		from django.http import HttpResponse
		from .middleware import SecurityHeadersMiddleware

		mw = SecurityHeadersMiddleware(lambda req: HttpResponse('x'))
		return mw(RequestFactory().get(path))['Content-Security-Policy']

	def test_la_api_no_puede_cargar_ningun_recurso(self):
		csp = self._csp('/api/maniobras/')
		self.assertIn("default-src 'none'", csp)

	def test_admin_puede_cargar_sus_propios_recursos_y_el_qr(self):
		csp = self._csp('/admin/otp_totp/totpdevice/1/config/')
		self.assertIn("default-src 'self'", csp)
		self.assertIn("img-src 'self' data:", csp)   # sin esto no carga el QR
		self.assertNotIn("default-src 'none'", csp)

	def test_ambas_rutas_siguen_prohibiendo_el_embebido(self):
		for path in ('/api/maniobras/', '/admin/'):
			self.assertIn("frame-ancestors 'none'", self._csp(path))


class CtaPortTipoServicioTests(SimpleTestCase):
	"""El tipo de servicio ya NO se adivina del texto: lo manda el botón
	'Tipo de Servicio'. Estos casos cubren la resolución del tipo, el dedup del
	segundo par (A19/B19) y el fallback de los registros viejos."""

	def _celdas(self, **campos):
		datos = {'folio': 'TEST-1', 'formato': 'excel'}
		datos.update(campos)
		resp = _generar_pdf_cta_port(datos, _TEMPLATE_PATH, 'test.pdf', formato='excel')
		self.assertEqual(resp.status_code, 200, getattr(resp, 'data', None))
		ws = load_workbook(io.BytesIO(resp.content))['CTA PORT FRABA CONTAINER']
		return ws

	def test_full_segundo_par_igual_no_se_escribe(self):
		ws = self._celdas(tipo_servicio='full', tipo='20 - 20 / DC - DC',
		                  contenedor='CBHU4284545 - TLLU3548653')
		self.assertEqual(ws['A17'].value, 2)
		self.assertEqual(ws['A18'].value, '20')
		self.assertEqual(ws['B18'].value, 'DC')
		self.assertIsNone(ws['A19'].value)
		self.assertIsNone(ws['B19'].value)
		self.assertEqual(ws['C17'].value, 'CBHU4284545 - TLLU3548653')

	def test_full_segundo_par_distinto_se_escribe(self):
		ws = self._celdas(tipo_servicio='full', tipo='20 - 40 / DC - HC',
		                  contenedor='CBHU4284545 - TLLU3548653')
		self.assertEqual(ws['A17'].value, 2)
		self.assertEqual(ws['A19'].value, '40')
		self.assertEqual(ws['B19'].value, 'HC')

	def test_sencillo_con_contenedor_largo_sigue_siendo_sencillo(self):
		# La heurística vieja (>12 caracteres) habría dicho full: ya no manda.
		ws = self._celdas(tipo_servicio='sencillo', tipo='20 / DC',
		                  contenedor='CBHU42845451234')
		self.assertEqual(ws['A17'].value, 1)
		self.assertEqual(ws['B17'].value, 'CONTENEDOR')

	def test_registro_viejo_sin_tipo_servicio_usa_heuristica(self):
		ws = self._celdas(tipo='20 / DC', contenedor='CBHU4284545-TLLU3548653')
		self.assertEqual(ws['A17'].value, 2)
		self.assertEqual(ws['B17'].value, 'CONTENEDORES')

	def test_carga_suelta_conserva_sus_dos_pares(self):
		ws = self._celdas(tipo_servicio='carga_suelta',
		                  tipo='9 - 14 / PALLETS - CARTONES',
		                  contenedor='CARGA SUELTA')
		self.assertEqual(ws['A17'].value, '9')
		self.assertEqual(ws['B17'].value, 'PALLETS')
		self.assertEqual(ws['A18'].value, '14')
		self.assertEqual(ws['B18'].value, 'CARTONES')
		self.assertEqual(ws['C17'].value, 'CARGA SUELTA')

	def test_full_con_un_solo_par_cuenta_uno(self):
		# La carga es editable en el modal: un Full puede salir con un solo
		# contenedor (el otro viaja con otro operador). El conteo sigue a lo que
		# se imprime, así que aquí es 1 CONTENEDOR — nunca "1 CONTENEDORES".
		ws = self._celdas(tipo_servicio='full', tipo='40 / HC',
		                  contenedor='CBHU4284545')
		self.assertEqual(ws['A17'].value, 1)
		self.assertEqual(ws['B17'].value, 'CONTENEDOR')
		self.assertIsNone(ws['A19'].value)
