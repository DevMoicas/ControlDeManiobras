import base64
import re
from decimal import Decimal
from zoneinfo import ZoneInfo
import django_filters
from rest_framework import viewsets, mixins
from .models import TorreFolio, Tracto, Remolque, Chofer, Maniobra, Gasto, Vacio, Empleado, Patio, Cliente, Origen, Destino, FotoRegistro, MovimientoLocal, Transportista, Cargo, UnidadTercero, OperadorTercero, DispositivoConfianza, DIAS_CONFIANZA, Folio, CostoExtra, Pendiente, LETRAS_CICLO, BATCH_SIZE, FORMATO_CODIGO, START_NUMERO, TorreControl, ReporteViaje, CARGAS_EN_EL_PAPEL
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from .Serializers import ReporteViajeSerializer, TorreFolioSerializer, TractoSerializer, RemolqueSerializer, ChoferSerializer, ManiobraSerializer, GastoSerializer, VacioSerializer, EmpleadoSerializer, PatioSerializer, ClienteSerializer, OrigenSerializer, DestinoSerializer, MovimientoLocalSerializer, TransportistaSerializer, CargoSerializer, UnidadTerceroSerializer, OperadorTerceroSerializer, FolioSerializer, CostoExtraSerializer, PendienteSerializer, TorreControlSerializer
from rest_framework.throttling import UserRateThrottle, AnonRateThrottle
from .throttling import SondeoThrottle
from rest_framework_simplejwt.views import TokenObtainPairView
from .Serializers import CustomTokenObtainPairSerializer, DispositivoConfianzaSerializer
from . import confianza
from .db_context import get_db_alias
from rest_framework.filters import OrderingFilter, SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import ValidationError
import os
import subprocess
import tempfile
import logging
import io
from datetime import date
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.db.models import Case, CharField, Count, F, IntegerField, Max, Q, Value, When
from django.db.models.functions import Concat, Substr
from django.http import FileResponse, HttpResponse
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.styles import Alignment
from PIL import Image
from rest_framework.views import APIView

logger = logging.getLogger(__name__)


# ── Documentos de viaje ──────────────────────────────────────────────────────
_TEMPLATE_PATH = (
    settings.BASE_DIR / 'api' / 'documentos' / 'templates' / 'CTA_PTE_FORMATO.xlsx'
)
_TEMPLATE_PATH_TERCEROS = (
    settings.BASE_DIR / 'api' / 'documentos' / 'templates' / 'FORMATO_CTA_PTE_TERCEROS.xlsx'
)
# NOTA: el template ya trae su logo. openpyxl SOLO conserva las imágenes al guardar
# si Pillow está instalado — por eso Pillow es dependencia obligatoria aquí.

_MESES_ES = {
    1:  'ENERO',
    2:  'FEBRERO',
    3:  'MARZO',
    4:  'ABRIL',
    5:  'MAYO',
    6:  'JUNIO',
    7:  'JULIO',
    8:  'AGOSTO',
    9:  'SEPTIEMBRE',
    10: 'OCTUBRE',
    11: 'NOVIEMBRE',
    12: 'DICIEMBRE',
}

# ponytail: el binario varía por SO/PATH — se prueban en orden hasta encontrar uno.
# libreoffice/soffice cubren Linux y Windows-con-PATH; la ruta completa cubre la
# instalación estándar de Windows cuando no está en el PATH. Si en prod (Linux) el
# binario tiene otra ruta, agregarla aquí.
_LIBREOFFICE_BINS = [
    'libreoffice',
    'soffice',
    r'C:\Program Files\LibreOffice\program\soffice.exe',
]


def _mayus(valor):
    """Todo lo que se escribe al Excel va en MAYÚSCULAS. Números/None se dejan igual."""
    return valor.upper() if isinstance(valor, str) else valor


_CHARS_FORMULA = ('=', '+', '-', '@')


def _sin_formula(valor):
    """Anti-inyección de fórmulas en Excel (CWE-1236): si un texto empieza por
    = + - o @, Excel/LibreOffice lo interpretaría como fórmula. Se antepone una
    comilla (') para forzar que sea texto. Es no-op en números/None y en textos
    que no empiezan por esos caracteres, así que un documento con datos legítimos
    queda idéntico; solo cambia el input malicioso."""
    if isinstance(valor, str) and valor[:1] in _CHARS_FORMULA:
        return "'" + valor
    return valor


def _xlsx_a_pdf(xlsx_path: str, output_dir: str) -> str:
    """
    Convierte un archivo .xlsx a PDF usando LibreOffice headless.
    Devuelve la ruta al PDF generado.
    Lanza Exception si LibreOffice retorna error.
    """
    env = os.environ.copy()
    env['HOME'] = output_dir   # LibreOffice necesita HOME para su perfil temporal

    result = None
    for binario in _LIBREOFFICE_BINS:
        try:
            result = subprocess.run(
                [
                    binario,
                    '--headless',
                    '--norestore',
                    '--convert-to', 'pdf',
                    '--outdir', output_dir,
                    xlsx_path,
                ],
                capture_output=True,
                timeout=60,
                env=env,
            )
            break
        except FileNotFoundError:
            continue

    if result is None:
        raise Exception(
            "LibreOffice no está instalado en el servidor (se buscó: "
            f"{', '.join(_LIBREOFFICE_BINS)})."
        )

    if result.returncode != 0:
        raise Exception(
            f"LibreOffice falló (código {result.returncode}): "
            f"{result.stderr.decode('utf-8', errors='replace')}"
        )

    pdf_name = os.path.splitext(os.path.basename(xlsx_path))[0] + '.pdf'
    pdf_path = os.path.join(output_dir, pdf_name)

    if not os.path.exists(pdf_path):
        raise Exception("LibreOffice no generó el archivo PDF esperado.")

    return pdf_path


_XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _responder_documento(wb, tmp_dir: str, basename: str, formato: str):
    """
    Guarda el workbook y devuelve un HttpResponse.
    formato == 'excel' → devuelve el .xlsx directo (sin convertir).
    cualquier otro valor → convierte a PDF con LibreOffice y lo devuelve.
    """
    xlsx_tmp = os.path.join(tmp_dir, f'{basename}.xlsx')
    wb.save(xlsx_tmp)

    if formato == 'excel':
        with open(xlsx_tmp, 'rb') as f:
            data = f.read()
        resp = HttpResponse(data, content_type=_XLSX_MIME)
        resp['Content-Disposition'] = f'attachment; filename="{basename}.xlsx"'
        return resp

    pdf_path = _xlsx_a_pdf(xlsx_tmp, tmp_dir)
    with open(pdf_path, 'rb') as f:
        pdf_bytes = f.read()
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{basename}.pdf"'
    return resp


def _concat_placas_remolques(placas: str, remolque_1: str, remolque_2: str) -> str:
    """
    Construye la cadena concatenada de placas y remolques.
    Formato:
      - Sin remolques:    "ABC-123"
      - Un remolque:      "ABC-123 / DEF-456"
      - Dos remolques:    "ABC-123 / DEF-456, GHI-789"
    """
    remolques = [r.strip() for r in [remolque_1, remolque_2] if r and r.strip()]
    if remolques:
        return f"{placas.strip()} / {', '.join(remolques)}"
    return placas.strip()


def _dividir_por_guion(segmento: str) -> tuple[str, str]:
    """
    Divide un segmento de texto en (parte_1, parte_2) usando '-' como separador.
    Si no hay '-', parte_2 queda vacía.
    Ej: "40 - 20" -> ("40", "20")   |   "40" -> ("40", "")
    """
    if '-' in segmento:
        izquierda, derecha = segmento.split('-', 1)
        return izquierda.strip(), derecha.strip()
    return segmento.strip(), ''


def _parsear_tipo(tipo_raw: str) -> tuple[str, str, str, str]:
    """
    Parsea el campo 'tipo' de Maniobra. El frontend garantiza el formato
    'IZQUIERDA / DERECHA' para registros nuevos (la diagonal es obligatoria
    y no editable en ManiobrasPage). Cada lado puede además tener un '-'
    interno para representar dos valores (ej. viajes full con dos tipos
    de contenedor distintos, o carga suelta con dos tipos de bulto).

    Ejemplos:
      "40 / HC"                      -> ("40", "",   "HC", "")
      "40 - 20 / HC - DC"            -> ("40", "20", "HC", "DC")
      "7 / PALLETS"                  -> ("7",  "",   "PALLETS", "")
      "9 - 14 / PALLETS - CARTONES"  -> ("9",  "14", "PALLETS", "CARTONES")
      "" o cualquier texto sin '/'   -> ("", "", "", "")  (registro viejo sin
                                          migrar: se trata como vacío, no error)

    Devuelve (numero_1, numero_2, letra_1, letra_2).
    """
    if not tipo_raw or '/' not in tipo_raw:
        return '', '', '', ''
    parte_izquierda, parte_derecha = tipo_raw.split('/', 1)
    numero_1, numero_2 = _dividir_por_guion(parte_izquierda)
    letra_1, letra_2   = _dividir_por_guion(parte_derecha)
    return numero_1, numero_2, letra_1, letra_2


def _es_carga_suelta(contenedor: str) -> bool:
    """
    True si la columna Contenedor de la maniobra contiene la palabra
    'CARGA SUELTA' (comparación tipo 'contains', no exige coincidencia exacta).
    """
    return 'CARGA SUELTA' in (contenedor or '').upper()


def _sumar_peso(peso_raw):
    """
    El campo peso puede traer dos cantidades separadas por '-' o '/', con o sin
    espacios (ej. "23412 - 22000", "8376/12117", "8376 / 12117"; carga suelta o
    full con dos pesos). Devuelve la suma numérica de las partes. Un solo valor
    se devuelve como número; si nada es numérico devuelve '' (celda en blanco,
    sin error).
    """
    partes = [p.strip() for p in re.split(r'[-/]', str(peso_raw)) if p.strip()]
    total = 0.0
    encontrado = False
    for p in partes:
        try:
            total += float(p)
            encontrado = True
        except (ValueError, TypeError):
            continue
    return total if encontrado else ''


def _validar_operador_vigente(nombre_operador):
    """
    Verifica que el operador (por nombre) no tenga licencia vencida.
    Devuelve (es_valido: bool, mensaje_error: str | None).
    Si el nombre no coincide con ningún chofer registrado, se permite
    (solo se bloquea si SÍ se encuentra el chofer Y su licencia está vencida).
    """
    if not nombre_operador:
        return True, None

    chofer = Chofer.objects.filter(nombre=nombre_operador).first()
    if not chofer:
        return True, None

    if chofer.fecha_vencimiento_licencia and chofer.fecha_vencimiento_licencia < date.today():
        return False, (
            f"No se puede asignar a {nombre_operador}: "
            f"su licencia venció el {chofer.fecha_vencimiento_licencia.strftime('%d/%m/%Y')}."
        )

    return True, None


def _generar_pdf_cta_port(request_data, template_path, nombre_archivo_pdf: str, formato: str = 'pdf'):
    """
    Lógica compartida para generar el PDF de la hoja 'CTA PORT FRABA CONTAINER'.
    La usan tanto DocumentoCtaPortView (Fraba Container) como
    DocumentoCtaPortTercerosView (Terceros) — ambos templates comparten
    EXACTAMENTE el mismo mapeo de celdas para esta hoja, solo cambia el
    archivo de template de origen y el nombre del PDF de salida.

    Devuelve:
      - HttpResponse con el PDF en caso de éxito.
      - rest_framework Response con status de error en caso de fallo.
    """
    # ── Leer datos del body (en MAYÚSCULAS para el Excel) ──────────────────
    folio       = request_data.get('folio', '').strip().upper()
    ccp         = request_data.get('ccp', '').strip().upper()
    origen      = request_data.get('origen', '').strip().upper()
    destino     = request_data.get('destino', '').strip().upper()

    # Datos del cliente
    cliente_nombre    = request_data.get('cliente_nombre', '').strip().upper()
    cliente_domicilio = request_data.get('cliente_domicilio', '').strip().upper()
    cliente_colonia   = request_data.get('cliente_colonia', '').strip().upper()
    cliente_ciudad    = request_data.get('cliente_ciudad', '').strip().upper()

    # Datos de la carga (auto-llenados desde maniobra via folio)
    tipo       = request_data.get('tipo', '').strip().upper()
    peso_raw   = request_data.get('peso', '')
    contenedor = request_data.get('contenedor', '').strip().upper()
    pedimento  = request_data.get('pedimento', '').strip().upper()
    referencia = request_data.get('referencia', '').strip().upper()

    # Datos de texto libre
    descripcion = request_data.get('descripcion', '').strip().upper()
    clave_sat   = request_data.get('clave_sat', '').strip().upper()

    # Fecha de expedición (manual, solo FRABA container; dd/MM/yyyy)
    fecha_expedicion = request_data.get('fecha_expedicion', '').strip()

    # Conductor y placas
    operador   = request_data.get('operador', '').strip().upper()
    placas     = request_data.get('placas', '').strip().upper()
    remolque_1 = request_data.get('remolque_1', '').strip().upper()
    remolque_2 = request_data.get('remolque_2', '').strip().upper()

    # Anti-inyección de fórmulas (CWE-1236) en los textos que van a celdas.
    # Se excluyen tipo/peso/fecha_expedicion: se parsean como número o tokens.
    (folio, ccp, origen, destino, cliente_nombre, cliente_domicilio,
     cliente_colonia, cliente_ciudad, contenedor, pedimento, referencia,
     descripcion, clave_sat, operador, placas, remolque_1, remolque_2) = map(
        _sin_formula,
        (folio, ccp, origen, destino, cliente_nombre, cliente_domicilio,
         cliente_colonia, cliente_ciudad, contenedor, pedimento, referencia,
         descripcion, clave_sat, operador, placas, remolque_1, remolque_2))

    # ── Validación mínima ──────────────────────────────────────────────────
    if not folio:
        return Response({'detail': 'El campo folio (Remisión) es requerido.'}, status=400)

    # ── Cálculos derivados ─────────────────────────────────────────────────
    remision_valor = f"{folio} / {ccp}" if ccp else folio

    # El tipo de servicio lo dictamina el botón "Tipo de Servicio" de la maniobra.
    # Los registros anteriores a ese campo llegan sin él: ahí se conserva la
    # heurística vieja (>12 caracteres = full, texto "CARGA SUELTA") para que sus
    # documentos salgan exactamente igual que antes del cambio.
    tipo_servicio = (request_data.get('tipo_servicio') or '').strip().lower()
    if tipo_servicio:
        es_carga_suelta = tipo_servicio == 'carga_suelta'
        es_full         = tipo_servicio == 'full'
    else:
        es_carga_suelta = _es_carga_suelta(contenedor)
        es_full         = len(contenedor) > 12
    numero_1, numero_2, letra_1, letra_2 = _parsear_tipo(tipo)

    # El conteo mira lo que se va a IMPRIMIR, no la etiqueta del servicio: los
    # campos de la carga son editables en el modal y un Full puede salir con un
    # solo contenedor (el otro viaja con otro operador). Si el tipo llega con un
    # solo par, es 1 CONTENEDOR — nunca "1 CONTENEDORES".
    # Acotado a quien manda tipo_servicio explícito: los registros anteriores a
    # ese campo entran en es_full por la heurística del contenedor largo y su
    # `tipo` puede traer un solo par, así que se les conserva el 2 de siempre.
    lleva_dos = es_full and (not tipo_servicio or numero_2 or letra_2)
    cantidad_label  = 2 if lleva_dos else 1
    tipo_label      = 'CONTENEDORES' if cantidad_label > 1 else 'CONTENEDOR'

    clave_sat_celda = f"CLAVE SAT:{clave_sat}" if clave_sat else ''

    peso_valor = _sumar_peso(peso_raw)

    concat_placas = _concat_placas_remolques(placas, remolque_1, remolque_2)

    # ── Abrir template y trabajar en directorio temporal ──────────────────
    try:
        with tempfile.TemporaryDirectory() as tmp_dir:
            wb = load_workbook(str(template_path))
            ws = wb['CTA PORT FRABA CONTAINER']

            # Remisión: J2 (celda superior del rango fusionado J2:J3)
            ws['J2'] = remision_valor

            # Fecha de expedición del documento: día numérico / mes en español / año
            _hoy = date.today()
            ws['H5'] = _hoy.day
            ws['I5'] = _MESES_ES[_hoy.month]
            ws['J5'] = _hoy.year

            # Origen y Destino
            ws['B6'] = origen
            ws['G6'] = destino

            # Datos del cliente
            ws['G7']  = cliente_nombre
            ws['G8']  = cliente_domicilio
            ws['G9']  = cliente_colonia
            ws['G10'] = cliente_ciudad

            # Tabla de bultos — full/sencillo vs. carga suelta
            if es_carga_suelta:
                # No se escribe cantidad (A17) ni "CONTENEDOR/CONTENEDORES" (B17):
                # esas celdas se reutilizan para la primera pareja número/letra.
                ws['A17'] = numero_1
                ws['B17'] = letra_1
                if numero_2:
                    ws['A18'] = numero_2
                if letra_2:
                    ws['B18'] = letra_2
            else:
                ws['A17'] = cantidad_label    # 1 (sencillo) o 2 (full)
                ws['B17'] = tipo_label        # 'CONTENEDOR' o 'CONTENEDORES'
                ws['A18'] = numero_1
                ws['B18'] = letra_1
                # Segundo par (full) solo si aporta información distinta a la del
                # primero: "20/DC" + "20/DC" es redundante → A19/B19 quedan vacías.
                if (numero_2 or letra_2) and (numero_2, letra_2) != (numero_1, letra_1):
                    ws['A19'] = numero_2
                    ws['B19'] = letra_2

            ws['C16'] = f'REFERENCIA: {referencia}' if referencia else ''   # Referencia
            ws['C17'] = contenedor        # No. de contenedor (o "CARGA SUELTA")
            ws['C18'] = f'PEDIMENTO: {pedimento}' if len(pedimento) >= 18 else pedimento  # 18+ chars: con etiqueta; incompleto: tal cual; vacío: vacío
            ws['F17'] = peso_valor        # Peso numérico
            if fecha_expedicion:
                # Solo el día en H5; el mes y el año ya van en otras celdas de la plantilla
                ws['H5'] = fecha_expedicion.split('/')[0].lstrip('0')   # día sin cero inicial

            # Descripción y Clave SAT
            ws['C20'] = descripcion
            ws['C21'] = clave_sat_celda

            # Conductor y Placas (valores directos para PDF autónomo)
            ws['C23'] = operador
            ws['F23'] = concat_placas

            # Eliminar las otras hojas para exportar solo CTA PORT
            for nombre_hoja in [n for n in wb.sheetnames if n != 'CTA PORT FRABA CONTAINER']:
                del wb[nombre_hoja]

            basename = os.path.splitext(nombre_archivo_pdf)[0]
            return _responder_documento(wb, tmp_dir, basename, formato)

    except FileNotFoundError:
        return Response(
            {'detail': 'No se encontró el template Excel. Contacte al administrador.'},
            status=500,
        )
    except subprocess.TimeoutExpired:
        return Response(
            {'detail': 'La conversión a PDF tardó demasiado. Intente de nuevo.'},
            status=500,
        )
    except Exception:
        logger.exception("Error inesperado al generar documento CTA PORT")
        return Response(
            {'detail': 'No se pudo generar el documento. Contacte al administrador.'},
            status=500,
        )


class AuditoriaMixin:
    """Rellena created_by/updated_by con el username autenticado en cada alta y
    edición. Los timestamps los pone Django (auto_now_add/auto_now). Los 4 campos
    son read_only en el serializer (editable=False), así que el cliente no puede
    falsearlos; aquí se setean vía serializer.save(), que sí los permite."""

    def perform_create(self, serializer):
        usuario = getattr(self.request.user, 'username', '') or ''
        serializer.save(created_by=usuario, updated_by=usuario)

    def perform_update(self, serializer):
        usuario = getattr(self.request.user, 'username', '') or ''
        serializer.save(updated_by=usuario)


# ── Refresco automatico: el reloj de una tabla ───────────────────────────────
# La mitad servidor de useAutoRefresco.js. El navegador pregunta cada 3 s "¿ha
# cambiado algo?" y solo pide datos cuando la respuesta cambia.
#
# Dos numeros y no uno: MAX(updated_at) detecta altas y ediciones, pero NO los
# borrados — la fila que se va no baja el maximo de las que quedan, asi que el
# reloj se quedaria clavado y la fila borrada seguiria en pantalla para siempre.
# El COUNT es lo que cierra ese hueco.
#
# Se eligio esto y no WebSockets porque la factura real de Azure es ~$16.55/mes
# y Redis costaria mas que el sistema entero; ver el analisis del 2026-08-26.
class CambiosMixin:
    """Anade `GET <recurso>/cambios/` y el filtro `?modificado_desde=`.

    El reloj se calcula sobre el queryset YA FILTRADO: quien mira Vacios
    pendientes solo se entera de lo que afecta a esa vista, y un vacio que pasa
    a entregado se nota como una baja en su count — que es justo lo que hay que
    repintar.
    """

    # ponytail: sin indice en updated_at, a proposito y medido (2026-08-26). El
    # COUNT obliga a recorrer la tabla entera, asi que Postgres IGNORA el indice
    # y hace Seq Scan igual: 0.147 ms con indice contra 0.189 ms sin el, ruido.
    # Con 414 maniobras creciendo a 87 al ano, 5.000 filas (0.7% de un core en
    # sondeo) quedan a decadas vista. El dia que duela, la salida NO es indexar:
    # es quitar el COUNT y detectar los borrados de otra forma.
    @action(detail=False, methods=['get'], url_path='cambios',
            throttle_classes=[SondeoThrottle])
    def cambios(self, request):
        reloj = self.filter_queryset(self.get_queryset()).aggregate(
            t=Max('updated_at'), n=Count('id'))
        return Response({
            # Cadena vacia y no null: el front compara con === y '' evita tener
            # que distinguir "sin fecha" de "primera vez".
            't': reloj['t'].isoformat() if reloj['t'] else '',
            'n': reloj['n'],
        })

    def get_queryset(self):
        """`?modificado_desde=<ISO>` acota a lo tocado despues de esa marca.

        Con esto el refresco pide UNA o dos filas en vez de sesenta, y la lista
        del navegador se actualiza por id sin perder el scroll ni las paginas ya
        cargadas.

        Las filas con updated_at NULL quedan fuera a proposito: son las
        anteriores a la auditoria (383 de 414 maniobras) y, por definicion, nadie
        las ha tocado desde entonces. En cuanto se editan, auto_now les pone
        fecha y entran solas.
        """
        qs = super().get_queryset()
        desde = self.request.query_params.get('modificado_desde')
        if not desde:
            return qs
        marca = parse_datetime(desde)
        if marca is None:
            # Sin esto, una marca ilegible devolveria la tabla entera como si
            # todo hubiera cambiado, y el refresco pasaria de 40 bytes a 60 filas
            # cada 3 segundos sin que nadie lo note.
            raise ValidationError({'modificado_desde': 'Marca de tiempo no valida.'})
        return qs.filter(updated_at__gt=marca)


# ── Gasto automatico al asignar el folio ─────────────────────────────────────
# Ver docs/planes/PLAN_GASTO_AUTOMATICO.md (rama main).
_TRANSPORTISTA_PROPIO = 'FRABA CONTAINER'


def _es_de_fraba(maniobra):
    """Si el viaje lo hace FRABA y no un tercero.

    Vacio cuenta como propio: es como lo interpreta ya todo el sistema (ver
    OperadorSelector en el frontend) y es lo que tienen 367 de las 409 maniobras.
    Se normalizan espacios y mayusculas porque `transportista` es texto escrito a
    mano; lo que no hace es adivinar variantes ("FRABA" a secas cuenta como
    tercero, y eso se arregla en el catalogo, no aqui).
    """
    return (maniobra.transportista or '').strip().upper() in ('', _TRANSPORTISTA_PROPIO)


def _crear_gasto_del_folio(maniobra, usuario):
    """Crea el gasto del viaje al asignarle su folio. Devuelve si lo creo.

    Solo para maniobras de FRABA, y solo si no hay gasto todavia: get_or_create
    se apoya en el UNIQUE de gastos.maniobra_id, que es lo que impide que dos
    peticiones simultaneas cuelen dos.

    Nace en ceros; los importes se capturan despues, como siempre. El folio, el
    operador y el destino NO se copian: GastoSerializer los lee de la maniobra
    enlazada, asi que se mantienen al dia solos.
    """
    if not _es_de_fraba(maniobra):
        return False
    _, creado = Gasto.objects.get_or_create(
        maniobra=maniobra,
        defaults={
            # str() y no .isoformat(): el modelo dice DateField pero la columna
            # real es TEXT (mismo caso documentado en folios_recientes). Sobre un
            # str, isoformat() revienta.
            'fecha_entrega_mercancia': str(maniobra.fecha_entrega_mercancia or ''),
            'created_by': usuario,
            'updated_by': usuario,
        },
    )
    return creado


def _sincronizar_fecha_entrega(maniobra, usuario):
    """Copia al gasto la FECHA DE ENTREGA de su maniobra.

    El gasto nace con la fecha que la maniobra tuviera en ese momento
    (_crear_gasto_del_folio) y hasta ahora se quedaba con ella para siempre. El
    caso normal es justo el malo: el folio se asigna ANTES de saber la fecha, asi
    que el gasto nacia vacio y habia que teclearla otra vez en Gastos. La
    maniobra es la unica que la sabe, asi que es la que manda.

    Se llama solo cuando la fecha CAMBIA, asi que editar cualquier otra cosa de
    la maniobra no pisa el gasto. Si la fecha se borra en la maniobra, tambien se
    borra aqui: son el mismo dato y dos valores distintos serian una mentira.

    str() y no .isoformat(): sobre un str, isoformat() revienta — mismo cuidado
    que en _crear_gasto_del_folio.
    """
    gasto = Gasto.objects.filter(maniobra=maniobra).first()
    if gasto is None:
        return False
    nueva = str(maniobra.fecha_entrega_mercancia or '')
    if (gasto.fecha_entrega_mercancia or '') == nueva:
        return False
    gasto.fecha_entrega_mercancia = nueva
    gasto.updated_by = usuario
    gasto.save()
    return True


# ── Vacios automaticos al asignar el folio ───────────────────────────────────
# Mismo disparo que el gasto: el contenedor que trae el viaje hay que devolverlo,
# asi que nace en Vacios en cuanto la maniobra tiene folio. Carga suelta no lleva
# contenedor y no da de alta nada. A diferencia del gasto, esto NO se limita a
# FRABA: el contenedor se devuelve lo mueva quien lo mueva (usuario, 2026-08-26).
_SEPARADOR_CARGA = re.compile(r'\s*[-/]\s*')


def _mitades(valor, valor2):
    """Las dos mitades de una carga, venga en el formato que venga.

    Desde la 0035 cada mitad tiene su columna; los registros anteriores guardan
    las dos dentro de la primera ("A - B", "A/B") y no hubo backfill. Es el
    leerPar() de utils/dobleValor.mjs recortado a lo que hace falta aqui: no
    devuelve el separador original porque en Vacios se escribe uno nuevo.
    """
    segundo = (valor2 or '').strip()
    if segundo:
        return (valor or '').strip(), segundo
    partes = _SEPARADOR_CARGA.split((valor or '').strip(), 1)
    return partes[0].strip(), (partes[1].strip() if len(partes) > 1 else '')


def _numeros_del_tipo(maniobra):
    """Solo los numeros del TIPO DE CARGA: "20 - 40 / DC - HC" -> ("20", "40").

    En Vacios el tipo se anota por el tamano ("40", "20 - 40"), sin las letras
    (usuario, 2026-08-26). El segundo sale de tipo_2 en los registros nuevos y
    del propio `tipo` en los viejos, igual que el contenedor.
    """
    numero_1, numero_2, _, _ = _parsear_tipo(maniobra.tipo or '')
    return numero_1, numero_2 or _parsear_tipo(maniobra.tipo_2 or '')[0]


def _unir(*partes):
    return ' - '.join(p for p in partes if p)


def _filas_de_vacios(maniobra):
    """(contenedor, tipo, operador) de cada vacio que deja esta maniobra.

    Un Full con un solo operador va en UNA fila con los dos contenedores; en
    cuanto se reparte entre dos, una fila por operador con lo suyo. El tipo no
    repite el numero: un Full de dos 40HC es "40", uno mixto "20 - 40".
    """
    if (maniobra.tipo_servicio or '').strip().lower() == 'carga_suelta':
        return []
    # Registros anteriores a tipo_servicio: la carga suelta se reconocia por el
    # texto del contenedor, y "CARGA SUELTA" no es un contenedor que devolver.
    if _es_carga_suelta(maniobra.contenedor):
        return []
    contenedor_1, contenedor_2 = _mitades(maniobra.contenedor, maniobra.contenedor_2)
    if not contenedor_1:
        return []
    numero_1, numero_2 = _numeros_del_tipo(maniobra)
    operador_1 = (maniobra.asignacion_operador_status or '').strip()
    operador_2 = (maniobra.operador_2 or '').strip()
    if contenedor_2 and operador_2:
        return [(contenedor_1, numero_1, operador_1),
                (contenedor_2, numero_2 or numero_1, operador_2)]
    tipo = numero_1 if numero_2 == numero_1 else _unir(numero_1, numero_2)
    return [(_unir(contenedor_1, contenedor_2), tipo, operador_1)]


def _vacio_pendiente(contenedor):
    """El vacio que ya esta esperando por ese contenedor, si lo hay.

    icontains y no exacto para reconocer tambien la fila combinada de un Full
    ("CONT1 - CONT2"). Solo entre los pendientes: el mismo contenedor vuelve a
    pasar meses despues y ese viaje si necesita su fila nueva. Es lo unico que
    impide duplicar, porque `vacios` no tiene columna que apunte a la maniobra.
    """
    return Vacio.objects.filter(status='pendiente', contenedor__icontains=contenedor).first()


def _separar_full_repartido(filas, usuario):
    """El Full dado de alta con un solo operador ocupa UNA fila con los dos
    contenedores. Cuando aparece el segundo operador, esa fila pasa a ser la del
    primero y la del segundo se crea despues.

    Solo toca la fila si de verdad lleva los dos contenedores dentro: una que
    alguien ya haya separado o editado a mano se queda como esta.
    """
    (contenedor_1, tipo_1, operador_1), (contenedor_2, _, _) = filas
    fila = _vacio_pendiente(contenedor_1)
    if fila is None or contenedor_2 not in (fila.contenedor or ''):
        return
    fila.contenedor      = contenedor_1
    fila.tipo_contenedor = tipo_1
    fila.operador        = operador_1
    fila.updated_by      = usuario
    fila.save()


def _crear_vacios_del_folio(maniobra, usuario):
    """Da de alta en Vacios los contenedores del viaje. Devuelve cuantos creo.

    Nacen en 'pendiente' —lo que hay que hacer con ellos esta por hacer— y con
    el operador del viaje ya puesto en OP DEL VIAJE. Lo demas (patio, fechas,
    coordinador) se captura en Vacios como hasta ahora.
    """
    filas = _filas_de_vacios(maniobra)
    if len(filas) == 2:
        _separar_full_repartido(filas, usuario)
    creados = 0
    for contenedor, tipo, operador in filas:
        if _vacio_pendiente(contenedor):
            continue
        Vacio.objects.create(
            contenedor=contenedor,
            tipo_contenedor=tipo,
            operador=operador,
            status='pendiente',
            created_by=usuario,
            updated_by=usuario,
        )
        creados += 1
    return creados


# ── Asignacion automatica del folio ──────────────────────────────────────────
# La columna ASIGNACION de la pagina de Folios dice quien lleva ese folio. Se
# escribia a mano; ahora la deriva la maniobra que lo tiene puesto.
#
#   FRABA (o sin transportista) -> las dos primeras palabras del operador.
#   Tercero                     -> "TERCERO <transportista>".
#
# Mismo criterio de "de quien es el viaje" que el gasto automatico (_es_de_fraba)
# para no tener dos definiciones de tercero que puedan discrepar.
def _asignacion_del_folio(maniobra, operador):
    """Lo que va escrito en el folio que lleva `operador`. "" si aun no se sabe.

    Recortado a 40, que es lo que admite Folio.asignacion: un nombre largo se
    corta en vez de reventar el guardado de la maniobra.
    """
    if not _es_de_fraba(maniobra):
        return ('TERCERO ' + (maniobra.transportista or '').strip())[:40]
    return ' '.join((operador or '').split()[:2])[:40]


def _sincronizar_asignacion_folios(maniobra, antes):
    """Escribe en cada folio quien lo lleva y limpia el que se acaba de soltar.

    `antes` trae folio/folio_2 como estaban en la base. Limpiar el anterior no
    es cosmetico: disponibles() da por ocupado todo folio con algo escrito en
    ASIGNACION, asi que un nombre olvidado lo retiraria del talonario para
    siempre aunque nadie lo estuviera usando.

    Corre en CADA guardado, no solo cuando el folio pasa de vacio a lleno: el
    caso normal es asignar el folio antes de saber quien lo llevara, y al
    capturar el operador o el transportista despues tiene que rellenarse solo.
    Con la maniobra sin folio no hace ni una consulta.

    El vinculo es el CODIGO y no una FK (maniobras es managed=False): un folio
    de texto libre que no este en el catalogo no casa con ninguna fila y el
    update no toca nada.
    """
    for campo, operador in (('folio',   maniobra.asignacion_operador_status),
                            ('folio_2', maniobra.operador_2)):
        actual = (getattr(maniobra, campo) or '').strip()
        previo = (antes.get(campo) or '').strip()
        if previo and previo != actual:
            Folio.objects.filter(codigo=previo).update(asignacion='')
        if actual:
            Folio.objects.filter(codigo=actual).update(
                asignacion=_asignacion_del_folio(maniobra, operador))


class TractoViewSet(viewsets.ModelViewSet):
    queryset = Tracto.objects.all().order_by('id')
    serializer_class = TractoSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

class RemolqueViewSet(viewsets.ModelViewSet):
    queryset = Remolque.objects.all().order_by('id')
    serializer_class = RemolqueSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

class ChoferViewSet(viewsets.ModelViewSet):
    queryset = Chofer.objects.all().order_by('id')
    serializer_class = ChoferSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

def filtro_status(valor):
    """Maniobras que tienen `valor` entre sus status.

    El campo guarda hasta 2 status separados por coma ("activo,quemada"), así que
    un filtro exacto dejaría fuera a esa maniobra. Se busca el valor como SEGMENTO
    completo del combo: las anclas (^|,) y (,|$) cubren de una sola vez los tres
    casos (va solo, va primero, va segundo) y evitan falsos positivos por
    coincidencia parcial si algún día un id fuera substring de otro.

    Vive suelto porque lo usan el filtro de la lista y el conteo de
    `resumen_status`: con una copia en cada sitio, un cambio en la regla se
    aplicaría solo a la mitad y la tabla y el panel dirían cosas distintas.
    """
    return Q(status__regex=rf"(^|,){re.escape(valor)}(,|$)")


class ManiobraFilter(django_filters.FilterSet):
    """Filtros de la lista de maniobras. La regla del status vive en
    `filtro_status` porque el panel de seguimientos cuenta con la misma."""
    status = django_filters.CharFilter(method="filter_status")
    # Bandera TERCERO (columna `tercero`): el front solo manda ?tercero=1 cuando
    # quiere ver únicamente los registros marcados, así que la presencia del
    # parámetro basta — devolvemos los que tienen la marca puesta (no NULL / no "").
    tercero = django_filters.CharFilter(method="filter_tercero")
    # "Sigue en piso": ni transportista ni operador. Basta con tener CUALQUIERA
    # de los dos para que ya haya quien la mueva y deje de estar esperando en el
    # puerto (decidido con el usuario, 2026-08-25). Alimenta el desglose de
    # PENDIENTES del panel de seguimientos.
    sin_asignar = django_filters.CharFilter(method="filter_sin_asignar")

    class Meta:
        model = Maniobra
        fields = ["status", "tercero", "sin_asignar"]

    def filter_status(self, queryset, name, value):
        return queryset.filter(filtro_status(value))

    def filter_tercero(self, queryset, name, value):
        return queryset.exclude(tercero__isnull=True).exclude(tercero="")

    def filter_sin_asignar(self, queryset, name, value):
        # Sin Trim: los dos campos salen de un desplegable (TransportistaSelector
        # y OperadorSelector guardan el nombre del catalogo), no de texto libre,
        # y hoy no hay ni una fila con solo espacios.
        return queryset.filter(
            (Q(transportista__isnull=True) | Q(transportista=""))
            & (Q(asignacion_operador_status__isnull=True)
               | Q(asignacion_operador_status=""))
        )


def _resolver_cliente(maniobra, clientes_por_nombre):
    """Cliente del catálogo al que apunta una maniobra.

    El FK manda: es lo único que distingue dos clientes homónimos (dos "YAZAKI"
    con distinta dirección). El mapa por nombre es el fallback para los folios
    creados antes del FK, que solo guardaron el texto — ahí la ambigüedad sigue
    (devuelve siempre el mismo) y se corrige al reeditar la maniobra.

    Vive fuera del ViewSet para poder probarlo sin BD: `maniobras` es
    managed=False, así que el runner de tests no crea esa tabla.
    """
    if maniobra.cliente_fk_id:
        return maniobra.cliente_fk
    return clientes_por_nombre.get(maniobra.cliente)


# --- NUEVA VISTA ---
class OrdenNullsLast(OrderingFilter):
    """OrderingFilter con las filas sin dato al final, ordene como ordene.

    Postgres pone los NULL PRIMERO en DESC, asi que ordenar por FECHA PIS
    descendente encabezaria la tabla con las maniobras a las que todavia no se
    les ha puesto (18 de 412 hoy): justo las que menos se estan mirando.
    """
    def filter_queryset(self, request, queryset, view):
        orden = self.get_ordering(request, queryset, view)
        if not orden:
            return queryset
        return queryset.order_by(*[
            F(campo[1:]).desc(nulls_last=True) if campo.startswith('-')
            else F(campo).asc(nulls_last=True)
            for campo in orden
        ])


class ManiobraViewSet(CambiosMixin, AuditoriaMixin, viewsets.ModelViewSet):
    # prefetch_related: la columna Costos Extra necesita los enlaces de cada
    # fila. Sin esto, una página de 60 maniobras hace 60 consultas extra.
    queryset = Maniobra.objects.all().prefetch_related('costos_extra_links').order_by("-id")
    serializer_class = ManiobraSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    filter_backends = [DjangoFilterBackend, OrdenNullsLast]
    filterset_class = ManiobraFilter
    ordering_fields = ["id", "fecha_pis", "fecha_entrega_mercancia", "sin_pis"]
    # La UNICA fecha que ordena esta tabla es FECHA PIS (decidido con el usuario
    # el 2026-08-25, tras probar a agrupar tambien por fecha de entrega: movia
    # servicios de sitio sin que se viera por que). Arriba las que no la tienen
    # —estan pendientes de que se la pongan— y debajo de la mas proxima hacia
    # atras. El id desempata: sin el, dos maniobras del mismo dia salen en orden
    # arbitrario y la paginacion de 60 en 60 puede repetir o saltarse filas.
    ordering = ["sin_pis", "-fecha_pis", "-id"]

    def get_queryset(self):
        """Anade `sin_pis`, un campo de orden virtual: 0 sin FECHA PIS, 1 con
        ella. Existe para poder pedirlo desde ?ordering= como un campo mas, en
        vez de imponer el criterio a todas las consultas de maniobras — el
        desglose de PENDIENTES pega al mismo endpoint y NO lo quiere.

        Va como campo agrupador y no apoyandose en nulls_first porque
        OrdenNullsLast manda los NULL al final: sin esto, una maniobra sin FECHA
        PIS cae al fondo de SU grupo y aparece a media tabla.

        Solo __isnull: la columna es DATE de verdad en Postgres (no TEXT como
        otras fechas de esta tabla), asi que "" no es un valor posible y
        compararlo reventaria la consulta.
        """
        return super().get_queryset().annotate(
            sin_pis=Case(
                When(fecha_pis__isnull=True, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            ),
        )

    def create(self, request, *args, **kwargs):
        es_valido, mensaje = _validar_operador_vigente(request.data.get('asignacion_operador_status', ''))
        if not es_valido:
            return Response({'detail': mensaje}, status=400)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        operador = request.data.get('asignacion_operador_status', None)
        if operador is not None:
            es_valido, mensaje = _validar_operador_vigente(operador)
            if not es_valido:
                return Response({'detail': mensaje}, status=400)
        return super().update(request, *args, **kwargs)

    # ── Gasto automatico ────────────────────────────────────────────────
    # Va aqui y no en el frontend aunque el folio se elija alli: se asigna desde
    # cuatro sitios de la pantalla de Maniobras (fila nueva, modal, celda de la
    # tabla y el vaciado al cambiar de plaza) y cada uno tendria que acordarse.
    # Aqui es un punto solo, y cubre cualquier via futura.
    #
    # atomic con las dos escrituras juntas: si la del gasto fallara, el folio
    # tampoco se asigna. Asi reintentar vuelve a disparar la regla — si la
    # maniobra quedara guardada, el folio ya no pasaria "de vacio a lleno" y el
    # gasto no se crearia nunca.

    def perform_create(self, serializer):
        with transaction.atomic(using=get_db_alias()):
            super().perform_create(serializer)
            maniobra = serializer.instance
            if (maniobra.folio or '').strip():
                _crear_gasto_del_folio(maniobra, self._usuario())
                _crear_vacios_del_folio(maniobra, self._usuario())
            _sincronizar_asignacion_folios(maniobra, {})

    def perform_update(self, serializer):
        # Los folios de ANTES: hasta el save(), serializer.instance trae la fila
        # como esta en la base.
        antes = {'folio':   serializer.instance.folio,
                 'folio_2': serializer.instance.folio_2}
        folio_antes = (antes['folio'] or '').strip()
        # Si el viaje ya tenia segundo operador antes de esta edicion: es lo que
        # distingue "se acaba de repartir el Full" de "ya venia repartido".
        operador_2_antes = (serializer.instance.operador_2 or '').strip()
        # La fecha de entrega de ANTES, por el mismo motivo: solo se propaga al
        # gasto cuando cambia de verdad.
        fecha_entrega_antes = serializer.instance.fecha_entrega_mercancia
        with transaction.atomic(using=get_db_alias()):
            super().perform_update(serializer)
            maniobra = serializer.instance
            # Solo al pasar de vacio a lleno. Cambiar un folio por otro no crea
            # nada: el gasto sigue enlazado a la maniobra y lee el folio de ella.
            if not folio_antes and (maniobra.folio or '').strip():
                _crear_gasto_del_folio(maniobra, self._usuario())
            # Los vacios se dan de alta al asignar el folio y, ademas, el dia que
            # aparece el segundo operador: hasta entonces el Full vive en una
            # sola fila y ahi hay que partirla en dos. Acotado a esas dos
            # transiciones para que editar cualquier otra cosa de una maniobra
            # vieja no resucite vacios ya entregados.
            if (maniobra.folio or '').strip() and (
                    not folio_antes
                    or (not operador_2_antes and (maniobra.operador_2 or '').strip())):
                _crear_vacios_del_folio(maniobra, self._usuario())
            # La fecha de entrega baja al gasto cada vez que cambia en la
            # maniobra: se pone casi siempre DESPUES de asignar el folio, asi que
            # copiarla solo al crear el gasto lo dejaba vacio para siempre.
            if maniobra.fecha_entrega_mercancia != fecha_entrega_antes:
                _sincronizar_fecha_entrega(maniobra, self._usuario())
            # La asignacion, en cambio, se recalcula siempre: el folio suele
            # ponerse antes de saber quien lo llevara.
            _sincronizar_asignacion_folios(maniobra, antes)

    def _usuario(self):
        return getattr(self.request.user, 'username', '') or ''

    @action(detail=False, methods=['get'], url_path='resumen-status')
    def resumen_status(self, request):
        """Cuántas maniobras hay en cada status. Alimenta el panel SEGUIMIENTOS
        de la pantalla de inicio.

        Existe en vez de contarlo desde el cliente porque la lista va paginada de
        60 en 60 y `page_size` no es configurable: leer el `count` desde el front
        obligaría a traerse 60 registros completos por cada status para mostrar
        dos números. Aquí son dos COUNT(*) y dos enteros.

        Cuenta por segmento (ver filtro_status): una maniobra "por_salir,activo"
        cuenta como activa, igual que en el filtro de la tabla.
        """
        return Response({
            estado: Maniobra.objects.filter(filtro_status(estado)).count()
            for estado in ('activo', 'pendiente')
        })

    @action(detail=False, methods=['get'], url_path='folios-recientes')
    def folios_recientes(self, request):
        """Devuelve UNA FILA POR FOLIO de las últimas 30 maniobras, incluyendo
        unidad (placas/tipo/modelo del tracto) y remolques para autollenar los
        documentos a partir del folio elegido.

        Un Full repartido entre dos operadores gasta un folio por operador y
        produce DOS filas: cada una con el operador, el tracto, los remolques y
        el contenedor que le tocan. Así elegir el folio en el modal ya elige de
        quién es el documento — no hace falta un selector de operador aparte."""
        con_folio   = Q(folio__isnull=False)   & ~Q(folio='')
        con_folio_2 = Q(folio_2__isnull=False) & ~Q(folio_2='')
        consulta = Maniobra.objects.filter(con_folio | con_folio_2)

        # `?placas=` acota a los folios de UNA unidad. Lo usa la torre de control:
        # en la fila del NO. 01 solo tienen sentido los folios que hizo el NO. 01.
        #
        # El filtro va aquí y no en el navegador a propósito: el corte de 30 es lo
        # último que se aplica, así que filtrando después, una unidad que llevara
        # días sin salir se quedaría sin ningún folio que ofrecer. Filtrando antes,
        # cada unidad tiene sus 30.
        #
        # Sin el parámetro, el endpoint se comporta igual que siempre: es lo que
        # siguen usando los documentos.
        placas = (request.query_params.get('placas') or '').strip()
        if placas:
            consulta = consulta.filter(Q(unidad=placas) | Q(unidad_2=placas))

        maniobras = consulta.select_related('cliente_fk').order_by('-id')[:30]
        # Mapa placas → tracto: resuelve Tipo de Unidad y Modelo sin N consultas.
        # Incluye los dos tractos: la Bitácora de Sueño del operador 2 necesita
        # el suyo igual que la del 1.
        placas_unidad = {
            placas
            for m in maniobras
            for placas in (m.unidad, m.unidad_2)
            if placas
        }
        tractos = {
            t.placas: t
            for t in Tracto.objects.filter(placas__in=placas_unidad)
        } if placas_unidad else {}
        # Mapa nombre → cliente, mismo patrón que tractos: SOLO para los folios
        # previos al FK cliente_id, que únicamente guardaron el nombre (ver
        # _resolver_cliente). Un nombre que ya no exista en el catálogo devuelve la
        # ficha vacía y el usuario la completa con el ClienteSelector.
        nombres_cliente = {
            m.cliente for m in maniobras if m.cliente and not m.cliente_fk_id
        }
        clientes = {
            c.nombre_cliente: c
            for c in Cliente.objects.filter(nombre_cliente__in=nombres_cliente)
        } if nombres_cliente else {}

        data = []
        for m in maniobras:
            cliente = _resolver_cliente(m, clientes)
            # Dos operadores = el reparto ya está decidido en la maniobra, y cada
            # folio lleva lo suyo. El front usa esto para NO ofrecer el
            # desplegable 1/2/Los dos: el documento no debe poder contradecir lo
            # guardado.
            dos_operadores = bool(m.operador_2 or m.folio_2)

            comun = {
                'id':          m.id,
                'origen':      m.origen or '',
                'destino':     m.destino or '',
                'pedimento':   m.pedimento or '',
                # La CITA del reporte de viaje sale de estos dos juntos: el dia
                # de fecha_pis y la hora de horario. Se mandan por separado y los
                # une el frontend (citaDesdeManiobra en utils/reporteViaje.mjs),
                # porque juntarlos aqui exigiria decidir la zona horaria en el
                # servidor y el horario es la hora LOCAL a la que se capturo.
                #
                # str() y no .isoformat(): `maniobras` es managed=False y ya hubo
                # una columna que el modelo declaraba DateField siendo TEXT en la
                # base (ver fecha_entrega_mercancia, mas abajo). str() da
                # 'YYYY-MM-DD' con un date y deja la cadena igual si algun dia
                # resulta ser texto; isoformat() reventaria sobre un str y se
                # llevaria por delante todo el endpoint.
                'fecha_pis':   str(m.fecha_pis or ''),
                # Texto libre: en la base hay '14:00' y tambien '9:00'.
                'horario':     m.horario or '',
                'referencia':  m.referencia or '',
                'tipo_servicio': m.tipo_servicio or '',
                'dos_operadores': dos_operadores,
                # La carga viaja EN CRUDO, las dos columnas tal cual, y el reparto
                # lo hace cargaDeParte() en el front con `parte`. Es a propósito:
                # los registros anteriores a la 0035 guardan los dos contenedores
                # dentro de la primera columna ("20 - 40 / DC - HC") y partirlos
                # aquí exigiría duplicar en Python el partidor que ya vive —y ya
                # está probado— en utils/dobleValor.mjs. Dos copias de esa regla
                # es justo como se separan con el tiempo.
                'tipo':        m.tipo or '',
                'peso':        str(m.peso) if m.peso is not None else '',
                'contenedor':  m.contenedor or '',
                'tipo_2':       m.tipo_2 or '',
                'peso_2':       str(m.peso_2) if m.peso_2 is not None else '',
                'contenedor_2': m.contenedor_2 or '',
                # Fecha de entrega de la maniobra. OJO: el modelo dice DateField
                # pero la columna REAL es TEXT (la tabla es managed=False, así que
                # cambiar el modelo nunca alteró el esquema — ver schema.sql y la
                # 0001). Llega ya como 'YYYY-MM-DD', que es justo lo que espera el
                # front; str() la deja igual y seguiría valiendo si algún día la
                # columna pasara a ser date de verdad. Nada de .isoformat(): sobre
                # un str revienta y se lleva por delante todo el endpoint.
                # Vacía si la maniobra aún no la tiene: entonces el gasto la deja
                # en blanco y se elige a mano.
                'fecha_entrega_mercancia': str(m.fecha_entrega_mercancia or ''),
                # Cliente del registro (autollenado de la Carta Porte). El nombre
                # sale del catálogo cuando hay FK, así que un cliente renombrado
                # allí se refleja aquí; `m.cliente` es el respaldo de los viejos.
                # cliente_id viaja para que el ClienteSelector del modal sepa CUÁL
                # homónimo está puesto (si no, elegir el otro parece "el mismo").
                'cliente_id':        m.cliente_fk_id,
                'cliente_nombre':    (cliente.nombre_cliente if cliente else m.cliente) or '',
                'cliente_domicilio': (cliente.domicilio if cliente else '') or '',
                'cliente_colonia':   (cliente.colonia   if cliente else '') or '',
                'cliente_ciudad':    (cliente.ciudad    if cliente else '') or '',
                # Fechas del viaje. Van aquí para que la torre de control pueda
                # acomodar sus bolitas al elegir el folio, sin una segunda
                # petición. Vacías si la maniobra aún no las tiene — entonces la
                # torre no acomoda nada y las bolitas se colocan a mano.
                'ruta_inicio': m.ruta_inicio.isoformat() if m.ruta_inicio else '',
                'ruta_fin':    m.ruta_fin.isoformat() if m.ruta_fin else '',
            }

            if m.folio and (not placas or (m.unidad or '').strip() == placas):
                tracto = tractos.get(m.unidad)
                data.append({
                    **comun,
                    'folio':       m.folio,
                    # El CCP es de cada operador: la remisión del documento se
                    # compone como "folio / ccp", así que va emparejado al folio.
                    'ccp':         m.ccp or '',
                    # Operador, unidad y remolques del registro (autollenado de documentos)
                    'operador':    m.asignacion_operador_status or '',      # nombre del operador asignado
                    'placas':      m.unidad or '',                          # placas del tracto asignado
                    'tipo_unidad': (tracto.unidad if tracto else '') or '', # Tipo de Unidad
                    'anio':        str(tracto.anio) if tracto and tracto.anio is not None else '',  # Modelo
                    'remolque_1':  m.remolque or '',
                    'remolque_2':  m.remolque_2 or '',
                    # Con dos operadores este folio se queda SOLO con el primer
                    # contenedor; con uno solo se lleva los dos y el desplegable
                    # 1/2/Los dos decide (por defecto, los dos).
                    'parte': '1' if dos_operadores else 'ambos',
                })

            if m.folio_2 and (not placas or (m.unidad_2 or '').strip() == placas):
                tracto_2 = tractos.get(m.unidad_2)
                data.append({
                    **comun,
                    'folio':       m.folio_2,
                    'ccp':         m.ccp_2 or '',
                    'operador':    m.operador_2 or '',
                    'placas':      m.unidad_2 or '',
                    'tipo_unidad': (tracto_2.unidad if tracto_2 else '') or '',
                    'anio':        str(tracto_2.anio) if tracto_2 and tracto_2.anio is not None else '',
                    'remolque_1':  m.remolque_3 or '',
                    'remolque_2':  m.remolque_4 or '',
                    # El folio del segundo operador se lleva el segundo contenedor,
                    # esté guardado en su columna o todavía dentro de la primera.
                    'parte': '2',
                })
        return Response(data)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class CustomTokenObtainPairView(TokenObtainPairView):

    # Devuelve access + refresh con 'role' y 'username' en el payload.
    serializer_class = CustomTokenObtainPairSerializer

    def post(self, request, *args, **kwargs):
        # El serializer valida credenciales + MFA + equipo de confianza, y deja
        # dicho en sus atributos qué hacer con la cookie. Aquí, que es donde se
        # puede tocar la respuesta HTTP, se emite o se borra.
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)   # 401 si fallan credenciales/MFA
        respuesta = Response(serializer.validated_data, status=status.HTTP_200_OK)

        token = getattr(serializer, 'cookie_confianza_a_emitir', None)
        if token:
            respuesta.set_cookie(
                confianza.COOKIE_CONFIANZA, token,
                max_age=DIAS_CONFIANZA * 24 * 60 * 60,   # ventana absoluta, = a la de BD
                httponly=True,
                secure=not settings.DEBUG,               # en dev sobre http se puede probar
                samesite='Strict',                       # mismo origen (SWA + backend enlazado)
                path='/api/login/',                      # solo viaja al login (decisión 15)
            )

        return respuesta


def _cerrar_todas_las_sesiones(user):
    """Mete en la blacklist todos los refresh vigentes del usuario. No se puede
    apuntar solo a la sesión del equipo perdido —los refresh no están ligados al
    dispositivo— así que se cierran todas (decisión 13). Corre bajo el rol
    estándar: la 0018 le dio SELECT en outstanding + INSERT en blacklisted."""
    from django.utils import timezone
    from rest_framework_simplejwt.token_blacklist.models import (
        OutstandingToken, BlacklistedToken,
    )
    for ot in OutstandingToken.objects.filter(user=user, expires_at__gt=timezone.now()):
        BlacklistedToken.objects.get_or_create(token=ot)


class DispositivoConfianzaViewSet(viewsets.ViewSet):
    """Equipos de confianza del propio usuario: listar y revocar.

    El queryset se ata a request.user, así que nadie ve ni revoca los de otro
    (la restricción es aquí, no en RLS, igual que en toda la app — decisión A1).
    El alta NO está aquí: un equipo se marca de confianza al iniciar sesión con
    la casilla, nunca por este endpoint.
    """
    permission_classes = [IsAuthenticated]

    def _propios_vigentes(self, request):
        from django.utils import timezone
        return DispositivoConfianza.objects.filter(
            usuario=request.user,
            revocado_en__isnull=True,
            expira_en__gt=timezone.now(),
        )

    def list(self, request):
        datos = DispositivoConfianzaSerializer(
            self._propios_vigentes(request), many=True
        ).data
        return Response(datos)

    @action(detail=True, methods=['post'])
    def revocar(self, request, pk=None):
        from django.utils import timezone
        disp = self._propios_vigentes(request).filter(pk=pk).first()
        if disp is None:
            return Response({'detail': 'No encontrado.'},
                            status=status.HTTP_404_NOT_FOUND)
        disp.revocado_en = timezone.now()
        disp.save(update_fields=['revocado_en'])
        _cerrar_todas_las_sesiones(request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)
class GastoViewSet(viewsets.ModelViewSet):
    # select_related evita el N+1 que dispara GastoSerializer.folio
    # (source='maniobra.folio') al serializar cada fila. El orden real lo pone
    # get_queryset(); este order_by es solo la red de seguridad de DRF.
    queryset = Gasto.objects.select_related('maniobra').all().order_by('-id')
    serializer_class = GastoSerializer

    def get_queryset(self):
        """Los gastos, de la fecha de entrega mas nueva a la mas vieja.

        `gastos.fecha_entrega_mercancia` NO es una fecha en la base: es varchar
        y hoy convive en DOS formatos, el ISO que escribe el DatePicker
        ('2026-08-29') y el 'DD/MM/YYYY' que quedo de cuando se tecleaba a mano.
        Ordenar la columna tal cual seria ordenar texto: '21/06/2026' saldria
        despues de '20/10/2026' —manda el dia— y los dos formatos se
        intercalarian entre si. Por eso se ordena por una clave normalizada a
        ISO, que en texto ya ordena como fecha.

        Lo que no case con ninguno de los dos formatos ('' , '200', textos
        sueltos) vale NULL y cae al final con nulls_last: sin fecha no hay sitio
        donde colocarlo, y arriba estorbaria justo lo que se viene a mirar.

        Se normaliza al LEER y no se toca el dato guardado: reescribir la columna
        cambiaria lo que ve el usuario en filas que nadie pidio tocar. El id
        desempata para que la paginacion de 60 en 60 no repita ni se salte filas.

        ponytail: la clave se calcula al vuelo en cada consulta, sin indice.
        Son cientos de filas; si algun dia son cientos de miles, el camino es
        normalizar la columna a DATE de una vez.
        """
        clave_fecha = Case(
            When(fecha_entrega_mercancia__regex=r'^\d{4}-\d{2}-\d{2}$',
                 then=F('fecha_entrega_mercancia')),
            When(fecha_entrega_mercancia__regex=r'^\d{2}/\d{2}/\d{4}$',
                 then=Concat(Substr('fecha_entrega_mercancia', 7, 4), Value('-'),
                             Substr('fecha_entrega_mercancia', 4, 2), Value('-'),
                             Substr('fecha_entrega_mercancia', 1, 2))),
            default=None,
            output_field=CharField(),
        )
        return (super().get_queryset()
                .annotate(fecha_orden=clave_fecha)
                .order_by(F('fecha_orden').desc(nulls_last=True), '-id'))

    def perform_create(self, serializer):
        maniobra_id = self.request.data.get('maniobra')
        if not maniobra_id:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'maniobra': 'Este campo es requerido.'})
        try:
            maniobra = Maniobra.objects.get(id=maniobra_id)
        except Maniobra.DoesNotExist:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'maniobra': f'No existe una maniobra con id {maniobra_id}.'})
        usuario = getattr(self.request.user, 'username', '') or ''
        serializer.save(maniobra=maniobra, created_by=usuario, updated_by=usuario)

    def perform_update(self, serializer):
        # En update no tocamos la maniobra, solo los campos del gasto
        usuario = getattr(self.request.user, 'username', '') or ''
        serializer.save(updated_by=usuario)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

class VacioViewSet(CambiosMixin, AuditoriaMixin, viewsets.ModelViewSet):
    queryset = Vacio.objects.all().order_by("-id")
    serializer_class = VacioSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    # ?status=pendiente|entregado filtra en el backend (la página de Vacíos tiene
    # scroll infinito paginado: filtrar en cliente solo cubriría lo ya cargado).
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    # `reprogramado` alimenta la vista REPROGRAMADOS (?reprogramado=true). Va como
    # filtro propio y no como un valor de ?status porque es un estado paralelo:
    # un vacío entregado puede estar reprogramado.
    filterset_fields = ["status", "reprogramado"]
    ordering_fields = ["id"]
    ordering = ["-id"]

    def create(self, request, *args, **kwargs):
        es_valido, mensaje = _validar_operador_vigente(request.data.get('operador', ''))
        if not es_valido:
            return Response({'detail': mensaje}, status=400)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        operador = request.data.get('operador', None)
        if operador is not None:
            es_valido, mensaje = _validar_operador_vigente(operador)
            if not es_valido:
                return Response({'detail': mensaje}, status=400)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

class EmpleadoViewSet(viewsets.ModelViewSet):
    queryset = Empleado.objects.all().order_by("-id")
    serializer_class = EmpleadoSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    filter_backends = [OrderingFilter]
    ordering_fields = ["id"]
    ordering = ["id"]

    def get_queryset(self):
        """?cargo= filtra por cargo (lo usa el selector de Coordinador en Vacíos).

        Mismo patrón que UnidadTerceroViewSet con ?transportista. En el servidor y
        no en el cliente porque la lista va paginada: filtrando en el navegador
        solo se mirarían los 60 primeros empleados.

        iexact y no exact: el cargo se guarda como texto en `empleados` (tabla
        managed=False) y una mayúscula de más escrita a mano dejaría al empleado
        fuera del desplegable sin que nadie entienda por qué.
        """
        cargo = self.request.query_params.get('cargo')
        qs = super().get_queryset()
        return qs.filter(cargo__iexact=cargo) if cargo else qs

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

class PatioViewSet(viewsets.ModelViewSet):
    queryset = Patio.objects.all()
    serializer_class = PatioSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    filter_backends = [OrderingFilter]
    ordering = ["nombre"]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    filter_backends = [OrderingFilter]
    ordering = ['nombre_cliente']

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class OrigenViewSet(viewsets.ModelViewSet):
    queryset = Origen.objects.all()
    serializer_class = OrigenSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    filter_backends = [OrderingFilter]
    ordering = ['ciudad']

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class DestinoViewSet(viewsets.ModelViewSet):
    queryset = Destino.objects.all()
    serializer_class = DestinoSerializer
    throttle_classes = [UserRateThrottle, AnonRateThrottle]
    filter_backends = [OrderingFilter]
    ordering = ['ciudad']

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class MovimientoLocalViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    queryset               = MovimientoLocal.objects.all()
    serializer_class       = MovimientoLocalSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]
    filter_backends        = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields       = ['status']
    search_fields          = ['operador', 'movimiento', 'unidad', 'contenedor']
    ordering_fields        = ['fecha', 'id']
    ordering               = ['-id']

    def create(self, request, *args, **kwargs):
        es_valido, mensaje = _validar_operador_vigente(request.data.get('operador', ''))
        if not es_valido:
            return Response({'detail': mensaje}, status=400)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        operador = request.data.get('operador', None)
        if operador is not None:
            es_valido, mensaje = _validar_operador_vigente(operador)
            if not es_valido:
                return Response({'detail': mensaje}, status=400)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar registros.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class TransportistaViewSet(viewsets.ModelViewSet):
    queryset               = Transportista.objects.all()
    serializer_class       = TransportistaSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar registros.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class CargoViewSet(viewsets.ModelViewSet):
    queryset               = Cargo.objects.all()
    serializer_class       = CargoSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar registros.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class UnidadTerceroViewSet(viewsets.ModelViewSet):
    queryset               = UnidadTercero.objects.all()
    serializer_class       = UnidadTerceroSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def get_queryset(self):
        t = self.request.query_params.get('transportista')
        return super().get_queryset().filter(transportista=t) if t else super().get_queryset()

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar registros.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class OperadorTerceroViewSet(viewsets.ModelViewSet):
    queryset               = OperadorTercero.objects.all()
    serializer_class       = OperadorTerceroSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def get_queryset(self):
        t = self.request.query_params.get('transportista')
        return super().get_queryset().filter(transportista=t) if t else super().get_queryset()

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar registros.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class PendienteViewSet(mixins.ListModelMixin,
                        mixins.CreateModelMixin,
                        mixins.UpdateModelMixin,
                        mixins.DestroyModelMixin,
                        viewsets.GenericViewSet):
    """Listas de pendientes de los cinco tableros.

    Se borran a mano y NO caducan solos (decidido con el usuario el 2026-08-25;
    antes se iban a las 28 horas y el borrado no existía para nadie). El DELETE
    lo puede usar cualquier usuario autenticado, sin candado de admin: un
    pendiente es una nota de una lista compartida, no un registro de negocio con
    historia que perder, igual que soltar una unidad en la torre de control.

    Sin `retrieve`: nadie pide un pendiente suelto, la página los trae todos.

    ponytail: el borrado es definitivo y esta tabla no lleva auditoría, así que
    no queda rastro de quién borró qué. Si algún día importa, es el mismo bloque
    de created_by/updated_by que ya llevan Maniobra, Gasto y Vacio.
    """
    queryset               = Pendiente.objects.all()
    serializer_class       = PendienteSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]
    # Sin paginar: son cinco tableros y el front los necesita TODOS de una vez
    # para repartirlos. Con el PAGE_SIZE=60 global, los tableros de abajo se
    # quedarían a medias sin avisar.
    pagination_class       = None


class TorreControlViewSet(viewsets.ModelViewSet):
    """Las bolitas ocupadas del tablero de la torre de control.

    Una fila es una unidad ocupada; la ausencia de fila es la unidad libre. Por
    eso las cuatro operaciones son de uso normal y ninguna está reservada:
    colocar es POST, mover es PATCH de `fecha` y soltar en UNIDADES LIBRES es
    DELETE. Aquí borrar no es destruir un registro de negocio —no hay historia
    que perder— sino devolver una unidad a disponible, así que no lleva el
    candado de admin de los catálogos: el tablero lo maneja quien opera.
    """
    queryset               = TorreControl.objects.select_related('tracto')
    serializer_class       = TorreControlSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]
    # Sin paginar: el tablero deduce las unidades libres restando las ocupadas,
    # así que una lista a medias inventaría unidades libres que no lo están. Con
    # el tope de la tabla (nº de tractos × 2) nunca se acerca al PAGE_SIZE=60.
    pagination_class       = None


class TorreFolioViewSet(mixins.ListModelMixin,
                        mixins.DestroyModelMixin,
                        viewsets.GenericViewSet):
    """Qué folio lleva cada unidad en la torre de control.

    Sin `update`: reasignar es un POST que sustituye. Y sin `retrieve`: la torre
    siempre las quiere todas de una vez para pintar su tabla.
    """
    queryset               = TorreFolio.objects.select_related('tracto')
    serializer_class       = TorreFolioSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]
    # Sin paginar: son como mucho tantas filas como tractos.
    pagination_class       = None

    def create(self, request, *args, **kwargs):
        """Asigna un folio a una unidad, sustituyendo el que tuviera.

        Es un upsert y no un "borra y crea" desde el frontend a propósito: dos
        peticiones dejarían una ventana en la que la unidad no tiene folio y
        otro usuario podría llevarse el suyo.
        """
        tracto_id = request.data.get('tracto')
        folio     = (request.data.get('folio') or '').strip()
        if not tracto_id or not folio:
            return Response({'detail': 'Faltan la unidad o el folio.'}, status=400)

        tracto = Tracto.objects.filter(id=tracto_id).first()
        if tracto is None:
            return Response({'detail': 'Unidad desconocida.'}, status=400)

        # Un folio asignado queda bloqueado para las demás unidades. El error
        # dice CUÁL lo tiene: "ya está asignado" a secas obliga a ir a buscarlo.
        ocupado = TorreFolio.objects.filter(folio=folio).exclude(tracto=tracto).first()
        if ocupado is not None:
            return Response(
                {'detail': f'El folio {folio} ya está asignado a {ocupado.tracto.no_eco}.'},
                status=400,
            )

        with transaction.atomic():
            TorreFolio.objects.filter(tracto=tracto).delete()
            asignacion = TorreFolio.objects.create(tracto=tracto, folio=folio)

        return Response(TorreFolioSerializer(asignacion).data, status=201)


class CostoExtraViewSet(viewsets.ModelViewSet):
    """Catálogo de costos extra (Finanzas → Costos extra).

    Borrado solo para admin, igual que el resto de catálogos. Doble candado: el
    403 de aquí y, por debajo, el rol de Postgres — los no-admin corren con el
    alias `standard`, que no tiene DELETE sobre esta tabla (migración 0038).
    """
    queryset               = CostoExtra.objects.all()
    serializer_class       = CostoExtraSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar registros.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class FolioViewSet(viewsets.ModelViewSet):
    queryset               = Folio.objects.all()
    serializer_class       = FolioSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]
    filter_backends        = [DjangoFilterBackend]
    # 'codigo' lo usa Maniobras para localizar el id del folio que va a renombrar
    # al marcar/desmarcar un Full: sin él habria que descargarse la tabla entera
    # (sin paginar, crece 14 filas por lote) para sacar un id.
    filterset_fields       = ['tabla', 'codigo']
    # Sin paginar a propósito: Folio crece sin límite (14 filas por lote, para
    # siempre) y el frontend necesita SIEMPRE la lista completa de una tabla
    # para reconstruir los lotes en columnas. Con PAGE_SIZE=60 global, el
    # quinto lote (70 filas) perdería las últimas 10 en silencio.
    pagination_class       = None

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar registros.'},
                            status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    def perform_update(self, serializer):
        """Renombrar un folio arrastra a la maniobra que lo tenía asignado.

        No hay FK entre Maniobra y Folio (la tabla `maniobras` es managed=False):
        el vínculo es el propio `codigo`, que es unique. Por eso al cambiarlo hay
        que reescribirlo también donde se usó, o la maniobra quedaría apuntando a
        un código que ya no existe. atomic() porque las dos escrituras son una
        sola operación; el alias sale del router (ver generar(), más abajo).
        """
        anterior = serializer.instance.codigo
        with transaction.atomic(using=get_db_alias()):
            folio = serializer.save()
            if folio.codigo != anterior:
                # La torre bloquea su folio con un `unique`, asi que hay que
                # mirar ANTES de escribir: si el codigo nuevo ya se lo tiene
                # asignado otra unidad, el UPDATE de abajo reventaria con un
                # IntegrityError que el usuario ve como un 500 sin motivo. El
                # atomic() deshace el save() de arriba, asi que el folio se queda
                # como estaba. Mismo mensaje que TorreFolioViewSet.create: dice
                # CUAL unidad lo tiene, o hay que ir a buscarla a mano.
                ocupado = TorreFolio.objects.filter(folio=folio.codigo).first()
                if ocupado is not None:
                    # values_list con JOIN se comeria una fila huerfana (la FK va
                    # con db_constraint=False), y entonces la colision pasaria
                    # desapercibida. Se resuelve el No. Eco aparte.
                    eco = (Tracto.objects.filter(pk=ocupado.tracto_id)
                           .values_list('no_eco', flat=True).first())
                    raise ValidationError({
                        'detail': f'No se puede renombrar a "{folio.codigo}": la torre '
                                  f'ya se lo tiene asignado a {eco or "otra unidad"}. '
                                  f'Quitaselo en la torre y repite.'
                    })

                # Las DOS columnas: un Full repartido gasta un folio por operador
                # y el renombrado puede caer en cualquiera de ellas. Dejar fuera
                # folio_2 deja esa maniobra apuntando a un codigo que ya no
                # existe, y disponibles() vuelve a ofrecer el numero como libre.
                # updated_at a mano: auto_now solo corre en save(), no en un
                # update() masivo. Sin esto el renombrado seria invisible para el
                # refresco automatico (CambiosMixin lee ese reloj) y las demas
                # pantallas seguirian mostrando el codigo viejo hasta un F5.
                ahora = timezone.now()
                Maniobra.objects.filter(folio=anterior).update(
                    folio=folio.codigo, updated_at=ahora)
                Maniobra.objects.filter(folio_2=anterior).update(
                    folio_2=folio.codigo, updated_at=ahora)
                # Y la torre. Es el UNICO vinculo entre el tablero y Maniobras:
                # sin esto, la fila se queda con el codigo viejo, get_servicio()
                # no encuentra la maniobra y la unidad aparece con su folio y sin
                # viaje. Antes era raro —renombrar era manual—, pero con el "-2"
                # automatico de los Full pasa en cada maniobra que se marca.
                TorreFolio.objects.filter(folio=anterior).update(folio=folio.codigo)

    @action(detail=False, methods=['get'], url_path='disponibles')
    def disponibles(self, request):
        """Los 5 siguientes folios de `tabla` que ninguna maniobra está usando.

        Alimenta el desplegable de la columna FOLIO en Maniobras.
        """
        tabla = request.query_params.get('tabla')
        if tabla not in dict(Folio.TABLA_CHOICES):
            return Response({'detail': 'tabla inválida.'}, status=status.HTTP_400_BAD_REQUEST)

        # ponytail: "usado" se deriva de maniobras.folio/folio_2, sin columna de
        # estado — así vaciar o borrar una maniobra libera sus folios sola. Sin
        # índice en ninguna de las dos: es un scan por apertura del desplegable;
        # añadirlo si pesa.
        # Las DOS columnas cuentan: un Full repartido entre dos operadores gasta
        # un folio por operador, y omitir folio_2 aquí volvería a ofrecer como
        # libre uno ya asignado (mismo motivo que la validación del serializer).
        usados = [
            codigo
            for fila in Maniobra.objects.values_list('folio', 'folio_2')
            for codigo in fila
            if codigo
        ]
        # asignacion='' — un folio con algo escrito a mano en ASIGNACIÓN cuenta
        # como ocupado. Es la vía para cargar talonarios viejos (lotes anteriores)
        # y marcar los que ya se usaron fuera del sistema, sin columna de estado.
        # `asignacion` es blank/default='' y validate_asignacion normaliza null a
        # '', así que "vacío" siempre es '' y no hace falta mirar NULL.
        libres = (Folio.objects.filter(tabla=tabla, asignacion='')
                  .exclude(codigo__in=usados)
                  .order_by('numero')[:5])
        return Response(FolioSerializer(libres, many=True).data)

    @action(detail=False, methods=['post'], url_path='generar')
    def generar(self, request):
        """Genera un lote de 14 folios para `tabla`
        (body: {"tabla": "manzanillo"|"lazaro", "direccion": "anterior"?}).

        Sin `direccion` avanza: los 14 siguientes al número más alto. Con
        "anterior" retrocede: los 14 que preceden al más bajo, para cargar los
        talonarios que ya existían antes de que arrancara el sistema (Manzanillo
        empieza en 2279 y Lázaro en 323, sembrados por la migración 0031, así que
        sin esto todo lo anterior es inalcanzable).

        Las letras salen de LETRAS_CICLO por índice DENTRO del lote, así que un
        lote hacia atrás también empieza en F.

        Abierto a cualquier usuario autenticado — confirmado con el usuario,
        igual que el resto de altas/ediciones; solo destroy() se reserva a admin.
        """
        tabla = request.data.get('tabla')
        if tabla not in dict(Folio.TABLA_CHOICES):
            return Response({'detail': 'tabla inválida.'}, status=status.HTTP_400_BAD_REQUEST)
        hacia_atras = request.data.get('direccion') == 'anterior'

        # using=get_db_alias() NO es opcional: el router manda el ORM al alias
        # del hilo ('standard' en una petición normal), así que un atomic sobre
        # 'default' abriría la transacción en OTRA conexión y el
        # select_for_update de abajo correría en autocommit → 500. Es la misma
        # regresión que tumbó /api/login/ el 2026-07-23 (ver Serializers.py).
        with transaction.atomic(using=get_db_alias()):
            if hacia_atras:
                # Bloquea la fila MÁS BAJA, la que define dónde empieza el hueco
                # anterior — igual que la otra rama bloquea la más alta. Dos clics
                # simultáneos se serializan.
                primero = (Folio.objects.select_for_update()
                           .filter(tabla=tabla).order_by('numero').first())
                if primero is None:
                    return Response(
                        {'detail': 'No hay folios en esta tabla todavía: genera el primer lote antes de retroceder.'},
                        status=status.HTTP_400_BAD_REQUEST)
                siguiente = primero.numero - BATCH_SIZE
                if siguiente < 1:
                    return Response(
                        {'detail': 'No se puede retroceder más: el lote anterior caería por debajo de 1.'},
                        status=status.HTTP_400_BAD_REQUEST)
            else:
                # select_for_update sobre la última fila de esta tabla serializa
                # clics concurrentes de AÑADIR FOLIOS.
                ultimo = (Folio.objects.select_for_update()
                          .filter(tabla=tabla).order_by('-numero').first())
                # En producción la migración 0031 siembra el primer lote, así que
                # `ultimo` nunca es None; el fallback evita el 500 si alguien vacía
                # la tabla desde el admin (y es lo que hace arrancable el test, que
                # corre con las migraciones de `api` saltadas).
                siguiente = ultimo.numero + 1 if ultimo else START_NUMERO[tabla]
            formato = FORMATO_CODIGO[tabla]
            nuevos = [
                Folio(tabla=tabla, numero=siguiente + i, letra=letra,
                      codigo=formato.format(letra=letra, numero=siguiente + i))
                for i, letra in enumerate(LETRAS_CICLO)
            ]
            creados = Folio.objects.bulk_create(nuevos)

        return Response(FolioSerializer(creados, many=True).data,
                        status=status.HTTP_201_CREATED)


class FotoRegistroViewSet(viewsets.ViewSet):
    """
    ViewSet para subir, ver y eliminar fotos de maniobras y vacíos.
    Las fotos se almacenan como bytes en PostgreSQL (BinaryField).
    Responde con base64 data URIs para que el frontend las renderice directamente.

    Endpoints:
      GET  /api/fotos/?tipo=maniobra&registro_id=123    → foto_1 y foto_2 en base64
      POST /api/fotos/subir/   multipart                → sube foto al slot 1 o 2
      DELETE /api/fotos/eliminar/?tipo=...&registro_id=...&slot=1  → elimina foto
    """

    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    ALLOWED_MIMES = {'image/jpeg', 'image/png', 'image/webp'}
    MAX_SIZE      = 2 * 1024 * 1024  # 2 MB

    # ── Validadores de parámetros comunes ────────────────────────────────────

    def _validar_params(self, tipo, registro_id_raw, slot_raw=None):
        """
        Valida y convierte los parámetros comunes.
        Retorna (registro_id, slot, error_response) donde error_response es None si OK.
        """
        if tipo not in ('maniobra', 'vacio'):
            return None, None, Response(
                {'detail': 'tipo inválido. Use maniobra o vacio.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            registro_id = int(registro_id_raw)
        except (TypeError, ValueError):
            return None, None, Response(
                {'detail': 'registro_id debe ser un número entero.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        slot = None
        if slot_raw is not None:
            try:
                slot = int(slot_raw)
                if slot not in (1, 2):
                    raise ValueError
            except (TypeError, ValueError):
                return None, None, Response(
                    {'detail': 'slot debe ser 1 o 2.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        return registro_id, slot, None

    # ── GET /api/fotos/?tipo=maniobra&registro_id=123 ────────────────────────

    def list(self, request):
        tipo            = request.query_params.get('tipo', '')
        registro_id_raw = request.query_params.get('registro_id', '')

        registro_id, _, err = self._validar_params(tipo, registro_id_raw)
        if err:
            return err

        try:
            foto = FotoRegistro.objects.get(tipo=tipo, registro_id=registro_id)
        except FotoRegistro.DoesNotExist:
            return Response({'foto_1': None, 'foto_2': None})

        def a_base64(data, mime):
            if not data:
                return None
            b64 = base64.b64encode(bytes(data)).decode('utf-8')
            return f'data:{mime};base64,{b64}'

        return Response({
            'foto_1': a_base64(foto.foto_1, foto.foto_1_mime),
            'foto_2': a_base64(foto.foto_2, foto.foto_2_mime),
        })

    # ── POST /api/fotos/subir/ ───────────────────────────────────────────────

    @action(
        detail=False,
        methods=['post'],
        url_path='subir',
        parser_classes=[MultiPartParser, FormParser],
    )
    def subir(self, request):
        tipo            = request.data.get('tipo', '')
        registro_id_raw = request.data.get('registro_id', '')
        slot_raw        = request.data.get('slot', '')
        foto_file       = request.FILES.get('foto')

        registro_id, slot, err = self._validar_params(tipo, registro_id_raw, slot_raw)
        if err:
            return err

        # Evita fotos huérfanas: el registro (maniobra/vacío) debe existir.
        modelo = Maniobra if tipo == 'maniobra' else Vacio
        if not modelo.objects.filter(pk=registro_id).exists():
            return Response(
                {'detail': 'El registro indicado no existe.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not foto_file:
            return Response(
                {'detail': 'No se envió ningún archivo. Campo requerido: foto.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if foto_file.size > self.MAX_SIZE:
            return Response(
                {'detail': 'La imagen no puede superar 2 MB.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        foto_bytes = foto_file.read()

        # No confiar en content_type (lo fija el cliente): validar los bytes reales.
        try:
            imagen = Image.open(io.BytesIO(foto_bytes))
            formato_real = (imagen.format or '').upper()
            imagen.verify()
        except Exception:
            return Response(
                {'detail': 'El archivo no es una imagen válida.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        _FORMATO_A_MIME = {'JPEG': 'image/jpeg', 'PNG': 'image/png', 'WEBP': 'image/webp'}
        if formato_real not in _FORMATO_A_MIME:
            return Response(
                {'detail': 'Formato no permitido. Use JPG, PNG o WEBP.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        mime = _FORMATO_A_MIME[formato_real]

        registro, _ = FotoRegistro.objects.get_or_create(
            tipo=tipo,
            registro_id=registro_id,
        )

        if slot == 1:
            registro.foto_1      = foto_bytes
            registro.foto_1_mime = mime
        else:
            registro.foto_2      = foto_bytes
            registro.foto_2_mime = mime

        registro.save()

        return Response(
            {'detail': f'Foto {slot} guardada correctamente.'},
            status=status.HTTP_200_OK,
        )

    # ── DELETE /api/fotos/eliminar/?tipo=...&registro_id=...&slot=1 ──────────

    @action(
        detail=False,
        methods=['delete'],
        url_path='eliminar',
    )
    def eliminar(self, request):
        if not request.user.is_staff:
            return Response(
                {'detail': 'No tienes permisos para eliminar fotos.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        tipo            = request.query_params.get('tipo', '')
        registro_id_raw = request.query_params.get('registro_id', '')
        slot_raw        = request.query_params.get('slot', '')

        registro_id, slot, err = self._validar_params(tipo, registro_id_raw, slot_raw)
        if err:
            return err

        try:
            registro = FotoRegistro.objects.get(tipo=tipo, registro_id=registro_id)
        except FotoRegistro.DoesNotExist:
            return Response(
                {'detail': 'No hay fotos para este registro.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if slot == 1:
            registro.foto_1      = None
            registro.foto_1_mime = None
        else:
            registro.foto_2      = None
            registro.foto_2_mime = None

        # Si ambas fotos quedaron vacías, eliminar la fila completa
        if not registro.foto_1 and not registro.foto_2:
            registro.delete()
        else:
            registro.save()

        return Response({'detail': f'Foto {slot} eliminada correctamente.'})


class DocumentoBitacoraSuenoView(APIView):
    """
    POST /api/documentos/bitacora-sueno/
    Recibe datos del formulario, llena la hoja 'BITACORA DE SUEÑO'
    del template Excel y devuelve el PDF generado.
    Requiere autenticación JWT.
    Todos los valores de texto se escriben en MAYÚSCULAS.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def post(self, request):
        # ── Leer datos del body (en MAYÚSCULAS para el Excel) ──────────────────
        operador    = request.data.get('operador', '').strip().upper()
        placas      = request.data.get('placas', '').strip().upper()
        remolque_1  = request.data.get('remolque_1', '').strip().upper()
        remolque_2  = request.data.get('remolque_2', '').strip().upper()
        folio       = request.data.get('folio', '').strip().upper()
        unidad      = request.data.get('unidad', '').strip().upper()    # tracto.unidad (Tipo de Unidad)
        anio        = request.data.get('anio', '').strip().upper()      # tracto.anio   (Modelo)
        origen      = request.data.get('origen', '').strip().upper()
        destino     = request.data.get('destino', '').strip().upper()
        fecha_salida  = request.data.get('fecha_salida', '').strip().upper()   # DD/MM/YYYY
        fecha_llegada = request.data.get('fecha_llegada', '').strip().upper()  # DD/MM/YYYY
        formato       = request.data.get('formato', 'pdf')

        # Anti-inyección de fórmulas (CWE-1236) en los textos que van a celdas.
        (operador, placas, remolque_1, remolque_2, folio, unidad, anio,
         origen, destino, fecha_salida, fecha_llegada) = map(_sin_formula, (
            operador, placas, remolque_1, remolque_2, folio, unidad, anio,
            origen, destino, fecha_salida, fecha_llegada))

        # ── Validación mínima ──────────────────────────────────────────────────
        if not placas:
            return Response({'detail': 'El campo placas es requerido.'}, status=400)

        # ── Construir valor concatenado (D9 y B11) ────────────────────────────
        concat_valor = _concat_placas_remolques(placas, remolque_1, remolque_2)

        # ── Abrir template y trabajar en directorio temporal ──────────────────
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                wb = load_workbook(str(_TEMPLATE_PATH))  # data_only=False → preserva fórmulas
                ws = wb['BITACORA DE SUEÑO']

                # Escribir valores en las celdas objetivo
                ws['L7']  = operador       # Nombre del operador
                ws['D9']  = concat_valor   # Vehículo / remolques (concatenado)
                ws['U9']  = folio          # CTA Porte (folio)
                ws['B11'] = concat_valor   # Placas (mismo concatenado — vinculado a CTA PORT via fórmula existente)
                ws['O11'] = unidad         # Tipo de Unidad
                ws['X11'] = anio           # Modelo (año)
                ws['E13'] = origen         # Origen del viaje
                ws['P13'] = destino        # Destino del viaje
                ws['E15'] = fecha_salida   # Fecha de salida (DD/MM/YYYY)
                ws['O15'] = fecha_llegada  # Fecha de llegada (DD/MM/YYYY)
                # NOTA: D33 tiene la fórmula '=L7' — se deja intacta.

                # Eliminar las otras hojas para exportar solo BITÁCORA DE SUEÑO
                for nombre_hoja in [n for n in wb.sheetnames if n != 'BITACORA DE SUEÑO']:
                    del wb[nombre_hoja]

                return _responder_documento(wb, tmp_dir, 'bitacora_sueno', formato)

        except FileNotFoundError:
            return Response(
                {'detail': 'No se encontró el template Excel. Contacte al administrador.'},
                status=500,
            )
        except subprocess.TimeoutExpired:
            return Response(
                {'detail': 'La conversión a PDF tardó demasiado. Intente de nuevo.'},
                status=500,
            )
        except Exception:
            logger.exception("Error inesperado al generar Bitácora de Sueño")
            return Response(
                {'detail': 'No se pudo generar el documento. Contacte al administrador.'},
                status=500,
            )


class DocumentoCtaPortView(APIView):
    """
    POST /api/documentos/cta-port/
    Recibe datos del formulario, llena la hoja 'CTA PORT FRABA CONTAINER'
    del template Excel y devuelve el PDF generado.
    Requiere autenticación JWT.
    Todos los valores de texto se escriben en MAYÚSCULAS.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def post(self, request):
        return _generar_pdf_cta_port(
            request.data, _TEMPLATE_PATH, 'cta_port.pdf',
            request.data.get('formato', 'pdf'),
        )


class DocumentoCtaPortTercerosView(APIView):
    """
    POST /api/documentos/cta-port-terceros/
    Genera el PDF de Carta Porte para transportistas Terceros, usando el
    template CTA_PTE_TERCEROS_FORMATO.xlsx. Comparte exactamente el mismo
    mapeo de celdas y la misma lógica de negocio que DocumentoCtaPortView
    (ver _generar_pdf_cta_port). No genera Bitácora de Gastos.
    Requiere autenticación JWT.
    Todos los valores de texto se escriben en MAYÚSCULAS.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def post(self, request):
        return _generar_pdf_cta_port(
            request.data, _TEMPLATE_PATH_TERCEROS, 'cta_port_terceros.pdf',
            request.data.get('formato', 'pdf'),
        )


class DocumentoBitacoraGastosView(APIView):
    """
    POST /api/documentos/bitacora-gastos/
    Recibe los mismos datos que DocumentoCtaPortView más `total_gastos`.
    Llena la hoja 'BITACORA GASTOS' del template con valores directos
    (anulando las fórmulas cruzadas que apuntan a CTA PORT FRABA CONTAINER).
    Devuelve el PDF de BITACORA GASTOS.
    Requiere autenticación JWT.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def post(self, request):
        # ── Leer datos del body ────────────────────────────────────────────────
        # Datos compartidos con CTA PORT (vienen del mismo formulario)
        operador   = request.data.get('operador',   '').strip()
        placas     = request.data.get('placas',     '').strip()
        remolque_1 = request.data.get('remolque_1', '').strip()
        remolque_2 = request.data.get('remolque_2', '').strip()
        peso_raw   = request.data.get('peso',       '')
        destino    = request.data.get('destino',    '').strip()

        # Campo exclusivo de BITACORA GASTOS
        total_gastos_raw = request.data.get('total_gastos', '')
        formato          = request.data.get('formato', 'pdf')

        # Anti-inyección de fórmulas (CWE-1236). peso/total_gastos se parsean como número.
        operador, placas, remolque_1, remolque_2, destino = map(
            _sin_formula, (operador, placas, remolque_1, remolque_2, destino))

        # ── Validación ────────────────────────────────────────────────────────
        if not total_gastos_raw and total_gastos_raw != 0:
            return Response(
                {'detail': 'El campo Total de Gastos es requerido.'},
                status=400,
            )

        try:
            total_gastos = float(str(total_gastos_raw).replace(',', ''))
        except (ValueError, TypeError):
            return Response(
                {'detail': 'Total de Gastos debe ser un número válido.'},
                status=400,
            )

        # ── Valores derivados ──────────────────────────────────────────────────
        # I2: mismo concatenado de placas/remolques que va en F23 de CTA PORT
        concat_placas = _concat_placas_remolques(placas, remolque_1, remolque_2)

        # I3: peso numérico (suma las cantidades separadas por '-', igual que
        # la celda F17 de CTA PORT). '' cuando no hay valor numérico → se
        # normaliza a 0.0 para la lógica de escritura de abajo.
        peso_valor = _sumar_peso(peso_raw) or 0.0

        # ── Abrir template y trabajar en directorio temporal ──────────────────
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                wb = load_workbook(str(_TEMPLATE_PATH))  # preserva fórmulas y formatos
                ws = wb['BITACORA GASTOS']

                # B2 (=TODAY()) — NO SE TOCA. LibreOffice resuelve la fórmula.

                # I2 (merged I2:L2): Unidad / Placas
                # Anula la fórmula ='CTA PORT FRABA CONTAINER'!F23
                ws['I2'] = concat_placas

                # B3 (merged B3:G3): Nombre del conductor
                # Anula la fórmula ='CTA PORT FRABA CONTAINER'!C23
                ws['B3'] = operador

                # I3 (merged I3:L3): Peso
                # Anula la fórmula =SUM('CTA PORT FRABA CONTAINER'!F17:F19)
                ws['I3'] = peso_valor if peso_valor != 0.0 else ''

                # B4 (merged B4:G4): Destino
                # Anula la fórmula ='CTA PORT FRABA CONTAINER'!G6
                ws['B4'] = destino

                # I4 (merged I4:L4): Total de Gastos
                # La celda ya tiene formato de dinero "$"#,##0 en el template.
                # openpyxl preserva el formato al cargar; solo se escribe el número.
                ws['I4'] = total_gastos

                # Eliminar las otras hojas para exportar solo BITACORA GASTOS
                for nombre_hoja in [n for n in wb.sheetnames if n != 'BITACORA GASTOS']:
                    del wb[nombre_hoja]

                return _responder_documento(wb, tmp_dir, 'bitacora_gastos', formato)

        except FileNotFoundError:
            return Response(
                {'detail': 'No se encontró el template Excel. Contacte al administrador.'},
                status=500,
            )
        except subprocess.TimeoutExpired:
            return Response(
                {'detail': 'La conversión a PDF tardó demasiado. Intente de nuevo.'},
                status=500,
            )
        except Exception:
            logger.exception("Error inesperado al generar Bitácora de Gastos")
            return Response(
                {'detail': 'No se pudo generar el documento. Contacte al administrador.'},
                status=500,
            )


class AlertasVencimientoView(APIView):
    """
    GET /api/alertas-vencimiento/
    Devuelve licencias de choferes y pólizas de tractos que vencen en los
    próximos 30 días, ordenadas por fecha ascendente. Visible para todos los
    usuarios autenticados sin distinción de rol.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]

    def get(self, request):
        from datetime import timedelta
        hoy = date.today()
        limite_licencia = hoy + timedelta(days=30)   # licencias: 1 mes
        limite_poliza   = hoy + timedelta(days=14)   # pólizas: 2 semanas

        choferes_por_vencer = Chofer.objects.filter(
            fecha_vencimiento_licencia__isnull=False,
            fecha_vencimiento_licencia__gte=hoy,
            fecha_vencimiento_licencia__lte=limite_licencia,
        ).values('nombre', 'fecha_vencimiento_licencia')

        tractos_por_vencer = Tracto.objects.filter(
            fecha_vencimiento_poliza__isnull=False,
            fecha_vencimiento_poliza__gte=hoy,
            fecha_vencimiento_poliza__lte=limite_poliza,
        ).values('unidad', 'anio', 'placas', 'fecha_vencimiento_poliza')

        alertas = []

        for c in choferes_por_vencer:
            alertas.append({
                'tipo':      'licencia',
                'nombre':    c['nombre'] or '(sin nombre)',
                'fecha':     c['fecha_vencimiento_licencia'].strftime('%d/%m/%Y'),
                'fecha_raw': c['fecha_vencimiento_licencia'].isoformat(),
            })

        for t in tractos_por_vencer:
            nombre_tracto = ' '.join(
                str(v) for v in (t['unidad'], t['anio'], t['placas']) if v
            ).strip()
            alertas.append({
                'tipo':      'poliza',
                'nombre':    nombre_tracto or '(sin datos)',
                'fecha':     t['fecha_vencimiento_poliza'].strftime('%d/%m/%Y'),
                'fecha_raw': t['fecha_vencimiento_poliza'].isoformat(),
            })

        alertas.sort(key=lambda a: a['fecha_raw'])

        return Response(alertas)



# ── Reporte de viaje: el documento ───────────────────────────────────────────
_TEMPLATE_REPORTE = (
    settings.BASE_DIR / 'api' / 'documentos' / 'templates' / 'REPORTE COORDINADORES.xlsx'
)

# Las horas se guardan en UTC (USE_TZ=True, TIME_ZONE='UTC') y el navegador las
# pinta en la del usuario. Este documento lo arma el SERVIDOR, así que hay que
# convertir aquí o el papel saldría seis horas corrido. tzdata está en
# requirements, así que ZoneInfo también funciona en Windows.
_ZONA_OPERACION = ZoneInfo('America/Mexico_City')


def _fecha_doc(valor):
    """Fecha suelta → DD/MM/YYYY. Cadena vacía si no hay."""
    return valor.strftime('%d/%m/%Y') if valor else ''


def _fecha_hora_doc(valor):
    """Instante → DD/MM/YYYY HH:MM en la hora de operación."""
    if not valor:
        return ''
    return timezone.localtime(valor, _ZONA_OPERACION).strftime('%d/%m/%Y %H:%M')


def _numero_doc(valor):
    """Los calculados llegan del serializer como cadena (ver get_total). En la
    hoja tienen que ser NÚMEROS o Excel no los suma."""
    return Decimal(valor) if valor not in (None, '') else ''


def _marcar_si_no(ws, celda, valor):
    """Escribe SI o NO encima de la etiqueta "SI / NO" que trae el papel.

    Sin contestar (None) la etiqueta se queda intacta, para rodearla a mano: es
    justo lo que distingue "no" de "todavía no se sabe" en un formato que se
    llena a lo largo de varios días.
    """
    if valor is not None:
        ws[celda] = 'SI' if valor else 'NO'


def _dejar_solo_la_elegida(ws, celda_si, celda_no, elegida_si):
    """Dos palabras impresas en celdas distintas (el SI/NO de estadías): se borra
    la que no aplica. Si no hay respuesta se dejan las dos, por el mismo motivo
    que en _marcar_si_no."""
    if elegida_si is not None:
        ws[celda_no if elegida_si else celda_si] = ''


def _marcar_casilla(ws, celda):
    """Una X en la casilla que sigue a la palabra impresa.

    RECOLECCIÓN EN PUERTO trae las dos opciones escritas (I3 PROPIO, K3 TERCERO)
    y una casilla detrás de cada una (J3, L3). Se marca la que toca y las dos
    palabras se quedan, que es como se llena el papel.
    """
    ws[celda] = 'X'


# La casilla de OPERADOR O TERCERO (I4:J4) mide unos 26.5 caracteres de ancho, y
# la plantilla le dio alto para DOS líneas (la fila 4 está a 30.75, el doble que
# las demás). Con el wrap_text que ya trae, un nombre normal cabe a tamaño
# completo repartido en dos renglones, que es como se lee mejor en el papel.
#
# ponytail: umbral fijo, medido sobre la plantilla — 2 líneas × ~26.5. Si alguien
# cambia el ancho de I/J o el alto de la fila 4, este número hay que rehacerlo.
_MAX_OPERADOR_EN_DOS_LINEAS = 53

# El doble del alto normal de la plantilla (16.5), o sea dos renglones.
_ALTO_FILA_OPERADOR = 33.0


def _encoger_para_caber(ws, celda):
    """Último recurso: encoge la letra hasta que el texto quepa.

    Solo para lo que ni siquiera entra en las dos líneas que tiene la casilla.
    Hace falta porque una celda COMBINADA no autoajusta su alto —ni en Excel ni
    en LibreOffice—, así que un texto más largo de la cuenta se corta en silencio
    en vez de empujar la fila.

    shrinkToFit es incompatible con wrap_text, así que hay que apagarlo: por eso
    no se aplica siempre, o los nombres normales saldrían en una sola línea
    diminuta pudiendo salir en dos a tamaño completo.
    """
    actual = ws[celda].alignment
    ws[celda].alignment = Alignment(
        horizontal=actual.horizontal, vertical=actual.vertical,
        wrap_text=False, shrink_to_fit=True,
    )


def _ajustar_a_una_hoja(ws):
    """Una sola hoja, horizontal.

    El formato son 12 columnas anchas: en vertical no cabe a lo ancho, así que
    LibreOffice lo parte y saca una segunda página con el pico derecho, ilegible
    y sin encabezados.
    """
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    # fitToWidth/fitToHeight no hacen NADA sin este interruptor: el escalado a
    # página vive en <sheetPr><pageSetUpPr fitToPage="1"/>, no en <pageSetup>.
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    # Sin área de impresión, cualquier celda con formato fuera del formato
    # arrastraría una página en blanco detrás.
    ws.print_area = 'A1:L25'
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)
    # Doble alto en la fila de OPERADOR O TERCERO: la casilla I4:J4 mide unos
    # 26.5 caracteres y los nombres completos no caben en una línea. Con el
    # wrap_text que trae la plantilla, dos líneas los meten a tamaño completo.
    #
    # Va aquí y no en el .xlsx a propósito: es un requisito del documento, y en
    # el archivo se pierde en cuanto alguien lo reedita o lo reemplaza. Aquí lo
    # sujeta una prueba.
    ws.row_dimensions[4].height = _ALTO_FILA_OPERADOR


def _llenar_reporte_viaje(ws, reporte, datos):
    """Vuelca un ReporteViaje sobre la plantilla del papel.

    El mapa de celdas está en docs/planes/PLAN_REPORTE_COORDINADORES.md. Las
    etiquetas viven en una celda y el valor en la siguiente; donde hay celdas
    combinadas se escribe SIEMPRE en la esquina superior izquierda, que es la
    única que openpyxl deja tocar.

    `datos` es el serializer ya resuelto: de ahí salen KM TOTALES, RENDIMIENTO y
    el TOTAL de cada carga. La plantilla no trae fórmulas y el serializer es el
    único dueño de esos cálculos — recalcularlos aquí sería una segunda copia.
    """
    def escribir(celda, valor):
        # MAYÚSCULAS y anti-inyección de fórmulas, igual que en las cartas porte.
        ws[celda] = _sin_formula(_mayus(valor))

    # ── Identificación ──
    escribir('C2', _fecha_doc(reporte.fecha))
    escribir('G2', reporte.coordinador)
    escribir('J2', reporte.folio)
    # 'carga_suelta' → "CARGA SUELTA": el guion bajo es del modelo, no del papel.
    escribir('B3', (reporte.servicio or '').replace('_', ' '))
    escribir('E3', reporte.cliente)
    if reporte.recoleccion:
        _marcar_casilla(ws, 'J3' if reporte.recoleccion == 'propio' else 'L3')
    escribir('B4', reporte.origen)
    escribir('E4', reporte.destino)
    escribir('I4', reporte.operador)
    # El wrap_text de la plantilla reparte el nombre en las dos líneas de la
    # casilla. Solo lo que no cabe ni así se encoge.
    if len(reporte.operador or '') > _MAX_OPERADOR_EN_DOS_LINEAS:
        _encoger_para_caber(ws, 'I4')
    escribir('D5', _fecha_hora_doc(reporte.cita))
    escribir('I5', _fecha_hora_doc(reporte.salida_puerto))
    escribir('D6', _fecha_hora_doc(reporte.inicio_pactado))
    escribir('I6', _fecha_hora_doc(reporte.salida_real))

    # ── Información del viaje ──
    # "UNIDAD Y PORTAS" es una sola celda para tracto y remolques: se concatena
    # con el mismo helper que las cartas porte, para que el papel se lea igual.
    escribir('D8', _concat_placas_remolques(
        reporte.unidad or '', reporte.remolque_1 or '', reporte.remolque_2 or ''))
    escribir('H8', reporte.km_inicial if reporte.km_inicial is not None else '')
    escribir('K8', datos.get('km_totales') if datos.get('km_totales') is not None else '')
    escribir('D9', reporte.operador)
    escribir('H9', reporte.km_final if reporte.km_final is not None else '')
    escribir('K9', _numero_doc(datos.get('rendimiento')))
    escribir('D10', _fecha_hora_doc(reporte.llegada_cliente))
    escribir('I10', _fecha_hora_doc(reporte.descarga))

    # ── En trayecto: los cinco renglones del papel, fila 12 = orden 1 ──
    # De la sexta carga en adelante NO se imprime: la fila 17 ya es el aceite y
    # escribir ahí lo pisaria. Se capturan y cuentan para el total del diesel y
    # el rendimiento, que es para lo que se pidieron (usuario, 2026-08-25).
    por_orden = {c['orden']: c for c in datos.get('cargas', [])}
    for orden in range(1, CARGAS_EN_EL_PAPEL + 1):
        fila = 11 + orden
        carga = por_orden.get(orden)
        if not carga:
            continue
        # Todo por _numero_doc: estos valores llegan del serializer y ahí los
        # decimales son CADENA. Escribirlos tal cual dejaría celdas de texto que
        # el Excel no suma ni formatea.
        escribir(f'C{fila}', _numero_doc(carga.get('litros_diesel')))
        escribir(f'E{fila}', _numero_doc(carga.get('precio_litro')))
        escribir(f'G{fila}', _numero_doc(carga.get('total')))
        escribir(f'I{fila}', _numero_doc(carga.get('litros_urea')))
        escribir(f'K{fila}', _numero_doc(carga.get('total_urea')))

    escribir('C17', reporte.litros_aceite if reporte.litros_aceite is not None else '')
    escribir('E17', reporte.precio_aceite if reporte.precio_aceite is not None else '')
    _marcar_si_no(ws, 'G17', reporte.reparacion)
    escribir('I17', reporte.reparacion_que)
    escribir('K17', reporte.reparacion_costo if reporte.reparacion_costo is not None else '')
    _marcar_si_no(ws, 'C18', reporte.rescate)
    escribir('G18', reporte.rescate_unidad)
    escribir('J18', reporte.rescate_operador)

    # ── Regreso ──
    escribir('E20', _fecha_hora_doc(reporte.llegada_manzanillo))
    _marcar_si_no(ws, 'J20', reporte.maniobra_vacio)
    escribir('C21', reporte.patio_entrega)
    escribir('E21', _fecha_hora_doc(reporte.cita_vacio))
    escribir('H21', reporte.unidad_vacio)
    escribir('J21', reporte.operador_vacio)
    _dejar_solo_la_elegida(ws, 'E22', 'F22', reporte.estadias)
    escribir('I22', reporte.estadias_horas if reporte.estadias_horas is not None else '')

    # ── Pie. F25:L25 (el espacio de la firma) se deja vacío a propósito: se
    # firma en papel. ──
    escribir('A24', reporte.comentarios)


class ReporteViajeViewSet(viewsets.ModelViewSet):
    """Reportes de viaje de los coordinadores. Uno por folio.

    Sin candado de rol para leer, crear ni editar: lo llena y lo ve cualquier
    usuario autenticado (decisión del usuario, 2026-08-24). `destroy` sí se
    reserva a admin, como en el resto del proyecto — un reporte es un documento
    que se firma, y borrarlo no es una corrección, es tirar el registro.

    Paginado (el PAGE_SIZE=60 global): a diferencia de Folios o Pendientes, esta
    tabla crece sin techo y la pantalla la recorre con buscador, no de una vez.

    `?folio=` para preguntar si un folio ya tiene reporte antes de abrir uno
    nuevo, y no descubrirlo con un 400 al guardar.
    """
    queryset               = ReporteViaje.objects.prefetch_related('cargas')
    serializer_class       = ReporteViajeSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    throttle_classes       = [UserRateThrottle, AnonRateThrottle]
    filter_backends        = [DjangoFilterBackend]
    filterset_fields       = ['folio']

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return Response({'detail': 'No tienes permisos para eliminar reportes.'},
                            status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get'], url_path='documento')
    def documento(self, request, pk=None):
        """GET /api/reportes-viaje/{id}/documento/?formato=excel|pdf

        Rellena la plantilla del papel con lo capturado y la devuelve para
        descargar. `formato=excel` entrega el .xlsx tal cual; cualquier otro
        valor lo pasa por LibreOffice y devuelve el PDF — el mismo camino que las
        cartas porte, con los mismos modos de fallo.

        GET y no POST como los demás documentos: aquí no se manda nada, se pide
        un reporte que ya está guardado.
        """
        reporte = self.get_object()
        datos   = self.get_serializer(reporte).data
        formato = request.query_params.get('formato', 'pdf')

        # El folio va en el nombre del archivo: se limpia por si un registro
        # viejo trae un folio de texto libre con barras, acentos o espacios.
        seguro = re.sub(r'[^A-Za-z0-9_-]+', '_', reporte.folio).strip('_') or 'reporte'

        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                wb = load_workbook(str(_TEMPLATE_REPORTE))
                # worksheets[0] y no wb['Table 1']: el nombre de la hoja es el que
                # traía el archivo original y renombrarla no debe tumbar esto.
                hoja = wb.worksheets[0]
                _llenar_reporte_viaje(hoja, reporte, datos)
                _ajustar_a_una_hoja(hoja)
                return _responder_documento(
                    wb, tmp_dir, f'reporte_viaje_{seguro}', formato)

        except FileNotFoundError:
            return Response(
                {'detail': 'No se encontró la plantilla del reporte. Contacte al administrador.'},
                status=500)
        except subprocess.TimeoutExpired:
            return Response(
                {'detail': 'La conversión a PDF tardó demasiado. Intente de nuevo.'},
                status=500)
        except Exception:
            logger.exception('Error inesperado al generar el reporte de viaje')
            return Response(
                {'detail': 'No se pudo generar el documento. Contacte al administrador.'},
                status=500)
